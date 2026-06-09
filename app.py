# -*- coding: utf-8 -*-
"""
🥔 土豆数据工具箱 - 让数据工作变得像挖土豆一样简单有趣
一个可爱风格的Streamlit数据处理工具集
包含：数据比对回填、数据拆分器、数据聚合器等功能
"""

# 依赖库检查
try:
    import dns.resolver
    DNS_AVAILABLE = True
except ImportError:
    DNS_AVAILABLE = False

import streamlit as st
import pandas as pd
import numpy as np
import ipaddress
import re
from io import BytesIO, StringIO
import time
import os
import zipfile
from datetime import datetime

# 页面配置 - 土豆主题
st.set_page_config(
    page_title="🥔 土豆数据工具箱",
    page_icon="🥔",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# 自定义CSS样式 - 可爱土豆风格
# ============================================
st.markdown("""
<style>
    /* ===== 整体背景和字体 ===== */
    @import url('https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800&display=swap');
    
    * {
        font-family: 'Nunito', 'PingFang SC', 'Microsoft YaHei', sans-serif !important;
    }
    
    /* 背景色 */
    .stApp {
        background: linear-gradient(135deg, #FFF8E7 0%, #FFE4C4 50%, #FFDAB9 100%);
    }
    
    /* 页面主容器 */
    .main .block-container {
        background: rgba(255, 255, 255, 0.9);
        border-radius: 20px;
        padding: 1.5rem;
        box-shadow: 0 4px 20px rgba(139, 69, 19, 0.1);
    }
    
    /* ===== 标题样式 ===== */
    .potato-header {
        text-align: center;
        padding: 1rem 0 0.5rem 0;
        margin-bottom: 0.5rem;
    }
    
    .potato-title {
        font-size: 2.2rem;
        font-weight: 800;
        color: #8B4513;
        margin-bottom: 0.3rem;
    }
    
    .potato-subtitle {
        font-size: 1rem;
        color: #D2691E;
        font-weight: 600;
    }
    
    /* ===== 土豆浮动装饰 ===== */
    @keyframes potato-float {
        0%, 100% {
            transform: translateY(0px);
        }
        50% {
            transform: translateY(-8px);
        }
    }
    
    .potato-decoration {
        text-align: center;
        margin: 0.8rem 0;
        color: #8B4513;
        font-size: 1.5rem;
        letter-spacing: 0.5rem;
        animation: potato-float 3s ease-in-out infinite;
    }
    
    /* ===== 卡片样式 ===== */
    .potato-card {
        background: linear-gradient(145deg, #FFFEF9 0%, #FFF5E6 100%);
        border-radius: 16px;
        padding: 1rem 1.2rem;
        box-shadow: 0 2px 10px rgba(139, 69, 19, 0.08);
        border: 2px solid #DEB887;
    }
    
    .potato-card-header {
        font-size: 1.1rem;
        font-weight: 700;
        color: #8B4513;
        margin-bottom: 0.8rem;
    }
    
    /* ===== 按钮样式 ===== */
    .stButton > button {
        background: linear-gradient(135deg, #FFA500 0%, #FF8C00 100%);
        color: white;
        border: none;
        border-radius: 50px;
        padding: 0.6rem 2rem;
        font-size: 1rem;
        font-weight: 700;
        box-shadow: 0 4px 15px rgba(255, 140, 0, 0.4);
    }
    
    .stButton > button:hover {
        background: linear-gradient(135deg, #FFB733 0%, #FFA500 100%);
    }
    
    .stDownloadButton > button {
        background: linear-gradient(135deg, #32CD32 0%, #228B22 100%);
        color: white;
        border: none;
        border-radius: 50px;
        padding: 0.8rem 2rem;
        font-size: 1.1rem;
        font-weight: 700;
    }
    
    /* ===== 指标卡片 ===== */
    .metric-card {
        background: linear-gradient(145deg, #FFFAF0 0%, #FFE4C4 100%);
        border-radius: 14px;
        padding: 1rem;
        text-align: center;
        border: 2px solid #F5DEB3;
    }
    
    .metric-label {
        font-size: 0.85rem;
        color: #8B4513;
        font-weight: 600;
        margin-bottom: 0.3rem;
    }
    
    .metric-value {
        font-size: 1.6rem;
        font-weight: 800;
        color: #D2691E;
    }
    
    /* ===== 进度条样式 ===== */
    .stProgress > div > div {
        background: linear-gradient(90deg, #FFA500, #FFD700);
        border-radius: 20px;
        height: 10px !important;
    }
    
    /* ===== 侧边栏样式 ===== */
    .sidebar .stSidebar {
        background: linear-gradient(180deg, #FFF8DC 0%, #FFE4C4 100%);
    }
    
    /* ===== 成功/警告/错误提示 ===== */
    .success-cute {
        padding: 0.8rem 1.2rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #98FB98 0%, #90EE90 100%);
        border: 2px solid #32CD32;
        color: #006400;
        font-weight: 600;
        margin: 0.5rem 0;
    }
    
    .warning-cute {
        padding: 0.8rem 1.2rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #FFFACD 0%, #FFE4B5 100%);
        border: 2px solid #FFD700;
        color: #8B4513;
        font-weight: 600;
    }
    
    .error-cute {
        padding: 0.8rem 1.2rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #FFB6C1 0%, #FFA0A0 100%);
        border: 2px solid #FF6B6B;
        color: #8B0000;
        font-weight: 600;
    }
    
    /* ===== 标签页样式 ===== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
        background: #FFF5E6;
        border-radius: 14px;
        padding: 6px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px;
        font-weight: 600;
        color: #8B4513;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #FFA500 0%, #FF8C00 100%) !important;
        color: white !important;
    }
    
    /* ===== 隐藏默认元素 ===== */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* ===== 分隔线 ===== */
    hr {
        border: none;
        height: 2px;
        background: linear-gradient(90deg, transparent, #DEB887, transparent);
        margin: 1rem 0;
    }
    
    /* ===== 展开器样式 ===== */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #FFF8DC 0%, #FFE4C4 100%);
        border-radius: 12px;
        font-weight: 600;
        color: #8B4513;
    }
    
    /* ===== Footer ===== */
    .footer {
        text-align: center;
        padding: 0.8rem;
        color: #8B4513;
        font-size: 0.9rem;
    }
    
    /* ===== 工具卡片样式 ===== */
    .tool-card {
        background: linear-gradient(145deg, #FFFAF0 0%, #FFE4C4 100%);
        border-radius: 16px;
        padding: 1.5rem;
        border: 2px solid #DEB887;
        text-align: center;
        transition: transform 0.2s;
    }
    
    .tool-card:hover {
        transform: translateY(-2px);
    }
    
    .tool-icon {
        font-size: 3rem;
        margin-bottom: 0.5rem;
    }
    
    .tool-title {
        font-size: 1.2rem;
        font-weight: 700;
        color: #8B4513;
        margin-bottom: 0.3rem;
    }
    
    .tool-desc {
        font-size: 0.9rem;
        color: #D2691E;
    }
    
    /* ===== 响应式调整 ===== */
    @media (max-width: 768px) {
        .potato-title {
            font-size: 1.8rem;
        }
        .metric-value {
            font-size: 1.3rem;
        }
    }
</style>
""", unsafe_allow_html=True)


# ============================================
# 通用工具函数
# ============================================
def load_csv_file(file) -> pd.DataFrame:
    """加载CSV文件，自动检测编码"""
    encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']
    
    for encoding in encodings:
        try:
            df = pd.read_csv(file, encoding=encoding)
            if len(df.columns) > 0:
                return df
        except (UnicodeDecodeError, Exception):
            file.seek(0)
            continue
    
    # 如果所有编码都失败，尝试用 errors='replace'
    file.seek(0)
    try:
        return pd.read_csv(file, encoding='utf-8', errors='replace')
    except Exception as e:
        st.error(f"❌ CSV文件加载失败: {str(e)}")
        return None


def load_data_file(file) -> pd.DataFrame:
    """统一加载Excel和CSV文件"""
    file_name = file.name.lower()
    
    if file_name.endswith(('.xlsx', '.xls')):
        try:
            return pd.read_excel(file, engine='openpyxl' if file_name.endswith('.xlsx') else 'xlrd')
        except Exception as e:
            st.error(f"❌ Excel文件加载失败: {str(e)}")
            return None
    elif file_name.endswith('.csv'):
        return load_csv_file(file)
    else:
        st.error(f"❌ 不支持的文件格式: {file_name}")
        return None


def csv_to_bytes(df: pd.DataFrame, filename: str = "result.csv") -> bytes:
    """将DataFrame转换为CSV字节流用于下载"""
    output = StringIO()
    df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    return output.getvalue()


def excel_to_bytes(df: pd.DataFrame, filename: str = "result.xlsx") -> bytes:
    """将DataFrame转换为Excel字节流用于下载"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='数据')
    output.seek(0)
    return output.getvalue()


def compare_columns(df: pd.DataFrame, col1: str, col2: str) -> list:
    """比对两列数据，返回差异行索引列表
    
    Args:
        df: 数据DataFrame
        col1: 第一列名称
        col2: 第二列名称
    
    Returns:
        差异行索引列表（重置后的行号，从0开始）
    """
    diff_indices = []
    
    # 验证列名是否存在
    if col1 not in df.columns:
        raise KeyError(f"列 '{col1}' 不存在于数据中")
    if col2 not in df.columns:
        raise KeyError(f"列 '{col2}' 不存在于数据中")
    
    # 重置索引确保连续
    df_reset = df.reset_index(drop=True).copy()
    
    for idx in range(len(df_reset)):
        val1 = str(df_reset.loc[idx, col1]).strip() if pd.notna(df_reset.loc[idx, col1]) else ''
        val2 = str(df_reset.loc[idx, col2]).strip() if pd.notna(df_reset.loc[idx, col2]) else ''
        if val1 != val2:
            diff_indices.append(idx)
    return diff_indices


def export_with_highlight(df: pd.DataFrame, diff_indices: list, filename: str = "diff_result.xlsx") -> bytes:
    """导出带高亮的Excel文件
    
    Args:
        df: 数据DataFrame
        diff_indices: 差异行索引列表
        filename: 文件名
    
    Returns:
        Excel文件字节流
    """
    from openpyxl.styles import PatternFill
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='数据比对结果')
    
    output.seek(0)
    
    # 使用 openpyxl 打开并设置样式
    from openpyxl import load_workbook
    
    wb = load_workbook(output)
    ws = wb.active
    
    # 差异行用浅红色背景 (#FFCCCC)
    highlight_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # 获取差异行号（DataFrame索引 + 2，因为Excel有表头行）
    for idx in diff_indices:
        row_num = idx + 2  # +2 因为：1是表头，索引从0开始
        for cell in ws[row_num]:
            cell.fill = highlight_fill
    
    # 保存到新的BytesIO
    result = BytesIO()
    wb.save(result)
    result.seek(0)
    
    return result.getvalue()


def display_column_preview(df: pd.DataFrame):
    """显示列预览信息"""
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">📝 总行数</div>
            <div class="metric-value">{len(df):,}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">📊 总列数</div>
            <div class="metric-value">{len(df.columns)}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">❓ 空值数量</div>
            <div class="metric-value">{df.isnull().sum().sum()}</div>
        </div>
        """, unsafe_allow_html=True)


def excel_to_bytes_multi(dfs: list, base_filename: str = "data") -> bytes:
    """将多个DataFrame打包成zip文件"""
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for i, df in enumerate(dfs, 1):
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=f'数据_{i}')
            excel_buffer.seek(0)
            zip_file.writestr(f'{base_filename}_{i:04d}.xlsx', excel_buffer.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# ============================================
# IP处理工具函数
# ============================================
def parse_single_ip_segment(segment: str) -> tuple[list, str]:
    """解析单个IP段（不包含分隔符的单个片段）

    支持格式：
    - 单个IP：192.168.1.1
    - 范围格式：192.168.1.1-192.168.1.10
    - CIDR格式：192.168.1.0/24

    Args:
        segment: IP段字符串

    Returns:
        (IP列表, 错误信息) - 成功时错误信息为None
    """
    segment = segment.strip()
    if not segment:
        return [], None

    try:
        # CIDR格式
        if '/' in segment:
            network = ipaddress.ip_network(segment, strict=False)
            return [str(ip) for ip in network.hosts()], None

        # 范围格式
        if '-' in segment:
            parts = segment.split('-')
            if len(parts) != 2:
                return [], f"范围格式错误: {segment}"

            start_ip = parts[0].strip()
            end_ip = parts[1].strip()

            start = int(ipaddress.IPv4Address(start_ip))
            end = int(ipaddress.IPv4Address(end_ip))

            if start > end:
                return [], f"起始IP大于结束IP: {segment}"

            return [str(ipaddress.IPv4Address(ip)) for ip in range(start, end + 1)], None

        # 单个IP
        ipaddress.IPv4Address(segment)
        return [segment], None

    except ipaddress.AddressValueError as e:
        return [], f"IP格式无效: {segment}"
    except ValueError as e:
        return [], f"解析错误: {segment}"
    except Exception as e:
        return [], f"未知错误: {segment}"


def parse_ip_range(ip_str: str, separators: str = ",;；，\t\n") -> tuple[list, dict]:
    """解析IP段（支持混合格式），返回 (IP列表, 统计信息)

    支持格式（可混合使用）：
    - 单个IP：192.168.1.1
    - 范围格式：192.168.1.1-192.168.1.10
    - CIDR格式：192.168.1.0/24
    - 混合格式：192.168.1.1,192.168.1.5-192.168.1.10,192.168.2.0/24

    Args:
        ip_str: IP字符串
        separators: 分隔符集合，默认支持逗号、分号、制表符、换行等

    Returns:
        (IP列表, 统计信息字典)
        统计信息包含：
        - total_segments: 总片段数
        - success_segments: 成功解析的片段数
        - failed_segments: 失败片段数
        - failed_items: 失败的片段列表
        - errors: 错误信息列表
    """
    stats = {
        "total_segments": 0,
        "success_segments": 0,
        "failed_segments": 0,
        "failed_items": [],
        "errors": []
    }

    if pd.isna(ip_str) or not str(ip_str).strip():
        stats["errors"].append("空值")
        return [], stats

    ip_str = str(ip_str).strip()

    if not ip_str:
        stats["errors"].append("空值")
        return [], stats

    # 构建分隔符正则（将所有分隔符统一替换为逗号）
    separator_pattern = f"[{re.escape(separators)}]+"

    # 分割字符串
    segments = re.split(separator_pattern, ip_str)

    # 过滤空片段
    segments = [s.strip() for s in segments if s.strip()]

    stats["total_segments"] = len(segments)

    all_ips = []

    for segment in segments:
        ip_list, error = parse_single_ip_segment(segment)

        if error:
            stats["failed_segments"] += 1
            stats["failed_items"].append(segment)
            stats["errors"].append(error)
        else:
            stats["success_segments"] += 1
            all_ips.extend(ip_list)

    # 去重
    all_ips = list(dict.fromkeys(all_ips))  # 保持顺序去重

    return all_ips, stats


def aggregate_ips_continuous(ip_list: list) -> list:
    """聚合连续IP为IP段（连续模式）
    
    将连续的IP聚合成IP段，非连续的保留
    
    Args:
        ip_list: IP列表
    
    Returns:
        聚合后的IP列表（连续的成段，不连续的保留单个IP）
    """
    if not ip_list:
        return []
    
    # 转换为整数并排序
    ip_ints = sorted([int(ipaddress.IPv4Address(ip)) for ip in ip_list])
    
    ranges = []
    start = ip_ints[0]
    end = ip_ints[0]
    
    for ip in ip_ints[1:]:
        if ip == end + 1:
            end = ip
        else:
            if start == end:
                ranges.append(str(ipaddress.IPv4Address(start)))
            else:
                ranges.append(f"{ipaddress.IPv4Address(start)}-{ipaddress.IPv4Address(end)}")
            start = ip
            end = ip
    
    # 添加最后一个范围
    if start == end:
        ranges.append(str(ipaddress.IPv4Address(start)))
    else:
        ranges.append(f"{ipaddress.IPv4Address(start)}-{ipaddress.IPv4Address(end)}")
    
    return ranges


def aggregate_ips_mixed(ip_list: list) -> str:
    """聚合IP（混合模式：连续的成段，不连续的保留）
    
    Args:
        ip_list: IP列表
    
    Returns:
        用逗号分隔的聚合结果
    """
    ranges = aggregate_ips_continuous(ip_list)
    return ', '.join(ranges)


# ============================================
# 页面1：首页
# ============================================
def show_home():
    """显示首页"""
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🥔 土豆数据工具箱 🥔</h1>
        <p class="potato-subtitle">✨ 让数据工作变得像挖土豆一样简单有趣 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 欢迎语
    st.markdown("""
    <div class="potato-card" style="text-align: center; margin: 1rem 0;">
        <p style="font-size: 1.1rem; color: #8B4513; margin: 0;">
            👋 欢迎使用土豆数据工具箱！这里为您准备了各种实用的数据处理工具 🥔
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # 工具列表
    st.markdown("""
    <div class="potato-card" style="margin: 1.5rem 0;">
        <div class="potato-card-header">🛠️ 可用工具（点击进入使用）</div>
    </div>
    """, unsafe_allow_html=True)
    
    # 工具1：数据比对回填
    col1, col2 = st.columns(2, gap="large")
    
    with col1:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">🔄</div>
            <div class="tool-title">数据比对回填</div>
            <div class="tool-desc">将两个Excel文件按关键字段进行比对和回填</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传主表和数据源 → 选择匹配字段 → 自动回填
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_compare", use_container_width=True):
            st.session_state.page = "🔄 数据比对回填"
            st.rerun()
    
    with col2:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">✂️</div>
            <div class="tool-title">数据拆分器</div>
            <div class="tool-desc">将大型Excel文件按指定条数拆分成多个文件</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传Excel文件 → 设置拆分条数 → 一键拆分打包
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_split", use_container_width=True):
            st.session_state.page = "✂️ 数据拆分器"
            st.rerun()
    
    # 工具2：数据聚合器 + 域名提取器
    col3, col4 = st.columns(2, gap="large")
    
    with col3:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">🔗</div>
            <div class="tool-title">数据聚合器</div>
            <div class="tool-desc">将相同数据的行合并，让内容聚合更高效</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传Excel文件 → 选择聚合字段 → 一键合并
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_aggregate", use_container_width=True):
            st.session_state.page = "🔗 数据聚合器"
            st.rerun()
    
    with col4:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">🌐</div>
            <div class="tool-title">域名提取器</div>
            <div class="tool-desc">从URL中提取主域名或子域名</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传Excel文件 → 选择URL字段 → 自动提取域名
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_domain", use_container_width=True):
            st.session_state.page = "🌐 域名提取器"
            st.rerun()
    
    # 工具3：单位树构建器 + IP处理工具
    st.markdown("---")
    col5, col6 = st.columns(2, gap="large")
    
    with col5:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">🌳</div>
            <div class="tool-title">单位树构建器</div>
            <div class="tool-desc">根据单位数据自动构建组织架构树</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传数据 → 字段映射 → 自动分组与上级判定
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_unit_tree", use_container_width=True):
            st.session_state.page = "🌳 单位树构建器"
            st.rerun()
    
    with col6:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">🖥️</div>
            <div class="tool-title">IP处理工具</div>
            <div class="tool-desc">IP段拆分与聚合，支持CIDR和范围格式</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传数据 → 选择模式 → IP拆分/聚合
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_ip_tool", use_container_width=True):
            st.session_state.page = "🖥️ IP处理工具"
            st.rerun()
    
    # 工具4：数据差异行 + 域名解析工具
    st.markdown("---")
    col7, col8 = st.columns(2, gap="large")
    
    with col7:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">🔍</div>
            <div class="tool-title">数据差异行</div>
            <div class="tool-desc">逐行比对两列数据，快速找出差异</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传数据 → 选择比对列 → 高亮差异行
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_diff_tool", use_container_width=True):
            st.session_state.page = "🔍 数据差异行"
            st.rerun()
    
    with col8:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem;">
            <div class="tool-icon">🔮</div>
            <div class="tool-title">域名解析工具</div>
            <div class="tool-desc">查询域名的DNS解析记录（CNAME/A记录）</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                📁 上传文件 → 选择域名列 → 自动解析DNS
            </p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🚀 进入工具", key="go_dns_tool", use_container_width=True):
            st.session_state.page = "🔮 域名解析工具"
            st.rerun()
    
    # 更多工具
    st.markdown("---")
    col9, col10 = st.columns(2, gap="large")
    
    with col9:
        st.markdown("""
        <div class="tool-card" style="padding-bottom: 0.5rem; opacity: 0.7;">
            <div class="tool-icon">🚧</div>
            <div class="tool-title">更多工具...</div>
            <div class="tool-desc">更多实用工具正在开发中，敬请期待！</div>
            <p style="margin-top: 0.5rem; color: #8B4513; font-size: 0.85rem;">
                🥔 土豆正在努力种植新的工具...
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.button("🚀 敬请期待", key="go_more", use_container_width=True, disabled=True)
    
    # 版本更新
    st.markdown("""
    <div class="potato-card" style="margin: 1.5rem 0;">
        <div class="potato-card-header">📝 版本更新</div>
        
        <div style="margin-top: 1rem; color: #8B4513;">
            <p style="margin: 0.5rem 0; font-weight: 600;">
                <span style="background: linear-gradient(135deg, #FF6B6B, #FF4757); color: white; padding: 0.15rem 0.6rem; border-radius: 15px; font-size: 0.8rem; margin-right: 0.5rem;">🔮 v2.6</span>
                当前版本
            </p>
            <ul style="margin: 0.3rem 0; padding-left: 2rem; line-height: 1.8; font-size: 0.9rem;">
                <li>新增域名解析工具</li>
                <li>查询域名的DNS解析记录（CNAME/A记录）</li>
                <li>支持批量解析和进度显示</li>
                <li>统计解析成功率和结果预览</li>
            </ul>
            
            <p style="margin: 1rem 0 0.5rem 0; font-weight: 600;">
                <span style="background: linear-gradient(135deg, #FF6B6B, #FF4757); color: white; padding: 0.15rem 0.6rem; border-radius: 15px; font-size: 0.8rem; margin-right: 0.5rem;">🔍 v2.5</span>
            </p>
            <ul style="margin: 0.3rem 0; padding-left: 2rem; line-height: 1.8; font-size: 0.9rem;">
                <li>新增数据差异行工具</li>
                <li>支持逐行比对两列数据</li>
                <li>差异行用浅红色高亮显示</li>
                <li>导出带高亮的Excel文件</li>
            </ul>
            
            <p style="margin: 1rem 0 0.5rem 0; font-weight: 600;">
                <span style="background: linear-gradient(135deg, #FF6B6B, #FF4757); color: white; padding: 0.15rem 0.6rem; border-radius: 15px; font-size: 0.8rem; margin-right: 0.5rem;">🖥️ v2.4</span>
            </p>
            <ul style="margin: 0.3rem 0; padding-left: 2rem; line-height: 1.8; font-size: 0.9rem;">
                <li>新增IP处理工具功能</li>
                <li>支持IP段拆分（CIDR和范围格式）</li>
                <li>支持IP聚合（连续和混合模式）</li>
                <li>同一单位数据隔离处理</li>
            </ul>
            
            <p style="margin: 1rem 0 0.5rem 0; font-weight: 600;">
                <span style="background: linear-gradient(135deg, #32CD32, #228B22); color: white; padding: 0.15rem 0.6rem; border-radius: 15px; font-size: 0.8rem; margin-right: 0.5rem;">🌳 v2.3</span>
            </p>
            <ul style="margin: 0.3rem 0; padding-left: 2rem; line-height: 1.8; font-size: 0.9rem;">
                <li>新增单位树构建器功能</li>
                <li>支持10种分组自动判定</li>
                <li>智能上级节点判定</li>
                <li>支持按分组和区域筛选预览</li>
            </ul>
            
            <p style="margin: 1rem 0 0.5rem 0; font-weight: 600;">
                <span style="background: linear-gradient(135deg, #FFA500, #FF8C00); color: white; padding: 0.15rem 0.6rem; border-radius: 15px; font-size: 0.8rem; margin-right: 0.5rem;">🥔 v2.2</span>
                域名提取器版
            </p>
            <ul style="margin: 0.3rem 0; padding-left: 2rem; line-height: 1.8; font-size: 0.9rem;">
                <li>新增域名提取器功能</li>
                <li>支持政务类域名和普通域名</li>
                <li>支持提取主域名和子域名</li>
            </ul>
            
            <p style="margin: 1rem 0 0.5rem 0; font-weight: 600;">
                <span style="background: #D2691E; color: white; padding: 0.15rem 0.6rem; border-radius: 15px; font-size: 0.8rem; margin-right: 0.5rem;">🍠 v2.0</span>
                工具箱版
            </p>
            <ul style="margin: 0.3rem 0; padding-left: 2rem; line-height: 1.8; font-size: 0.9rem;">
                <li>重构为多工具集架构</li>
                <li>新增数据拆分器功能</li>
                <li>土豆主题UI优化</li>
            </ul>
            
            <p style="margin: 1rem 0 0.5rem 0; font-weight: 600;">
                <span style="background: #8B4513; color: white; padding: 0.15rem 0.6rem; border-radius: 15px; font-size: 0.8rem; margin-right: 0.5rem;">🥔 v1.0</span>
                初始版本
            </p>
            <ul style="margin: 0.3rem 0; padding-left: 2rem; line-height: 1.8; font-size: 0.9rem;">
                <li>数据比对回填功能</li>
                <li>可爱土豆风格界面</li>
            </ul>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 底部装饰
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="footer">
        <p>💡 选择上方工具开始使用吧！</p>
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 页面2：数据比对回填
# ============================================
def show_compare_tool():
    """显示数据比对回填工具"""
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🔄 数据比对回填</h1>
        <p class="potato-subtitle">✨ 将两个Excel文件按关键字段进行数据回填 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">将两个Excel文件按关键字段进行数据比对和回填，适合数据整合场景。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传主表 → ② 上传数据源 → ③ 选择匹配字段 → ④ 选择回填字段 → ⑤ 执行回填
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 初始化session state
    if 'df1' not in st.session_state:
        st.session_state.df1 = None
    if 'df2' not in st.session_state:
        st.session_state.df2 = None
    if 'result_df' not in st.session_state:
        st.session_state.result_df = None
    if 'stats' not in st.session_state:
        st.session_state.stats = None
    
    # 使用说明卡片
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0;">
            <span style="font-size: 2.5rem;">🥔</span>
            <h2 style="color: #8B4513; margin: 0.3rem 0;">使用说明</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="potato-card" style="margin-bottom: 0.8rem;">
            <div class="potato-card-header">🌱 操作步骤</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>上传 <b>主表</b> 📁（支持 Excel/CSV）</li>
                <li>上传 <b>数据源</b> 📁（支持 Excel/CSV）</li>
                <li>选择 <b>匹配字段</b> 🔍</li>
                <li>选择 <b>回填字段</b> 🔄</li>
                <li>点击 <b>开始比对</b> 🚀</li>
                <li>下载 <b>结果文件</b> 📥（Excel/CSV）</li>
            </ol>
        </div>
        
        <div class="potato-card">
            <div class="potato-card-header">💡 温馨提示</div>
            <ul style="color: #8B4513; line-height: 1.7; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>支持 .xlsx .xls .csv 格式</li>
                <li>CSV自动检测编码（UTF-8/GBK）</li>
                <li>大文件使用批量merge加速</li>
                <li>原始数据不会被修改</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 2rem;">🥔 🌿</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("🥔 数据比对回填")
    
    # 文件上传区域
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown('<div class="potato-card"><div class="potato-card-header">📁 主表（文件1）</div></div>', unsafe_allow_html=True)
        
        file1 = st.file_uploader(
            "点击上传或拖拽文件到此处",
            type=['xlsx', 'xls', 'csv'],
            help="🥔 主表将作为输出文件的基础（支持 .xlsx .xls .csv）",
            key="file_uploader_1"
        )
        
        if file1:
            with st.spinner("🥔 加载中..."):
                df1 = load_data_file(file1)
                if df1 is not None:
                    st.session_state.df1 = df1
                    # 显示文件信息
                    file_size_mb = file1.size / (1024 * 1024)
                    st.markdown(f"""
                    <div style="background: #E8F5E9; padding: 0.5rem 1rem; border-radius: 8px; margin: 0.5rem 0;">
                        <span style="color: #2E7D32; font-size: 0.85rem;">
                            📄 {file1.name} ({file_size_mb:.2f} MB) | 📝 {len(df1):,} 行 × {len(df1.columns)} 列
                        </span>
                    </div>
                    """, unsafe_allow_html=True)
                    st.markdown("""
                    <div class="success-cute">✅ 已加载主表文件</div>
                    """, unsafe_allow_html=True)
                    display_column_preview(df1)
    
    with col2:
        st.markdown('<div class="potato-card"><div class="potato-card-header">📁 数据源（文件2）</div></div>', unsafe_allow_html=True)
        
        file2 = st.file_uploader(
            "点击上传或拖拽文件到此处",
            type=['xlsx', 'xls', 'csv'],
            help="🍠 数据源提供要回填的数据（支持 .xlsx .xls .csv）",
            key="file_uploader_2"
        )
        
        if file2:
            with st.spinner("🍠 加载中..."):
                df2 = load_data_file(file2)
                if df2 is not None:
                    st.session_state.df2 = df2
                    # 显示文件信息
                    file_size_mb = file2.size / (1024 * 1024)
                    st.markdown(f"""
                    <div style="background: #E8F5E9; padding: 0.5rem 1rem; border-radius: 8px; margin: 0.5rem 0;">
                        <span style="color: #2E7D32; font-size: 0.85rem;">
                            📄 {file2.name} ({file_size_mb:.2f} MB) | 📝 {len(df2):,} 行 × {len(df2.columns)} 列
                        </span>
                    </div>
                    """, unsafe_allow_html=True)
                    st.markdown("""
                    <div class="success-cute">✅ 已加载数据源文件</div>
                    """, unsafe_allow_html=True)
                    display_column_preview(df2)
    
    # 数据预览
    if st.session_state.df1 is not None or st.session_state.df2 is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        
        preview_tab1, preview_tab2 = st.tabs(["📋 主表预览", "📋 数据源预览"])
        
        with preview_tab1:
            if st.session_state.df1 is not None:
                st.dataframe(st.session_state.df1.head(20), use_container_width=True, height=280)
            else:
                st.markdown("""
                <div style="text-align: center; padding: 2rem; color: #8B4513;">
                    🥔 请上传主表文件
                </div>
                """, unsafe_allow_html=True)
        
        with preview_tab2:
            if st.session_state.df2 is not None:
                st.dataframe(st.session_state.df2.head(20), use_container_width=True, height=280)
            else:
                st.markdown("""
                <div style="text-align: center; padding: 2rem; color: #8B4513;">
                    🍠 请上传数据源文件
                </div>
                """, unsafe_allow_html=True)
    
    # 字段配置区域
    if st.session_state.df1 is not None and st.session_state.df2 is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 字段配置</div></div>', unsafe_allow_html=True)
        
        # 字段数量对比
        col_count_info = st.columns([1, 2])
        with col_count_info[0]:
            st.markdown(f"""
            <div style="background: #FFF8DC; padding: 1rem; border-radius: 12px;">
                <div style="color: #8B4513; font-size: 0.9rem;">
                    <b>📊 字段数量</b>
                </div>
                <div style="margin-top: 0.5rem; color: #D2691E;">
                    主表：<b>{len(st.session_state.df1.columns)}</b> 个字段<br>
                    数据源：<b>{len(st.session_state.df2.columns)}</b> 个字段
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col_count_info[1]:
            st.markdown("""
            <div style="background: #FFE4C4; padding: 1rem; border-radius: 12px;">
                <div style="color: #8B4513; font-size: 0.9rem;">
                    <b>💡 说明</b>
                </div>
                <div style="margin-top: 0.5rem; color: #8B4513; font-size: 0.9rem;">
                    选择<b>匹配字段</b>和<b>回填字段</b>，回填字段会添加到结果表中（带<code>_来源</code>后缀）。
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<hr>", unsafe_allow_html=True)
        
        config_col1, config_col2, config_col3 = st.columns(3)
        
        with config_col1:
            match_col1 = st.selectbox(
                "🎯 主表匹配字段",
                options=["（请选择）"] + list(st.session_state.df1.columns),
                index=0,
                help="选择主表中用于匹配的字段"
            )
            if match_col1 == "（请选择）":
                match_col1 = ""
        
        with config_col2:
            match_col2 = st.selectbox(
                "🎯 数据源匹配字段",
                options=["（请选择）"] + list(st.session_state.df2.columns),
                index=0,
                help="选择数据源中用于匹配的字段"
            )
            if match_col2 == "（请选择）":
                match_col2 = ""
        
        with config_col3:
            fill_cols = st.multiselect(
                "🔄 回填字段",
                options=[col for col in st.session_state.df2.columns],
                default=[],
                help="选择要从数据源回填的字段"
            )
        
        # 字段预览
        if match_col1 and match_col2:
            preview_col1, preview_col2 = st.columns(2)
            
            with preview_col1:
                st.markdown(f"**🥔 主表 `{match_col1}` 预览**")
                if match_col1 in st.session_state.df1.columns:
                    unique_count = st.session_state.df1[match_col1].nunique()
                    null_count = st.session_state.df1[match_col1].isnull().sum()
                    st.caption(f"唯一值：{unique_count:,} | 空值：{null_count:,}")
                    st.write(st.session_state.df1[match_col1].dropna().head(8).tolist())
            
            with preview_col2:
                st.markdown(f"**🍠 数据源 `{match_col2}` 预览**")
                if match_col2 in st.session_state.df2.columns:
                    unique_count = st.session_state.df2[match_col2].nunique()
                    null_count = st.session_state.df2[match_col2].isnull().sum()
                    st.caption(f"唯一值：{unique_count:,} | 空值：{null_count:,}")
                    st.write(st.session_state.df2[match_col2].dropna().head(8).tolist())
            
            # 预估结果字段
            if fill_cols:
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown("**📋 结果预估**")
                
                # 计算预估的字段列表
                all_cols1 = list(st.session_state.df1.columns)
                # 回填字段都加_来源后缀
                fill_cols_new = [f"{col}_来源" for col in fill_cols]
                result_cols_count = len(all_cols1) + len(fill_cols_new)
                
                st.markdown(f"""
                <div style="background: #FFF8DC; padding: 1rem; border-radius: 12px; color: #8B4513;">
                    <p style="margin: 0.3rem 0;">• 主表字段：<code>{', '.join(all_cols1)}</code>（{len(all_cols1)}个）</p>
                    <p style="margin: 0.3rem 0;">• 回填字段（加_来源后缀）：<code>{', '.join(fill_cols_new)}</code>（{len(fill_cols_new)}个）</p>
                    <p style="margin: 0.3rem 0;">• <b>结果表总计：{result_cols_count} 个字段</b></p>
                </div>
                """, unsafe_allow_html=True)
        
        # 执行按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        
        # 大文件警告
        total_rows = len(st.session_state.df1) + len(st.session_state.df2)
        is_large_file = total_rows > 100000
        
        if is_large_file:
            st.markdown(f"""
            <div style="background: #FFF3E0; padding: 1rem; border-radius: 12px; margin-bottom: 1rem; border-left: 4px solid #FF9800;">
                <div style="color: #E65100; font-weight: 600; margin-bottom: 0.3rem;">
                    ⚠️ 大文件检测到
                </div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    当前数据量：<b>{len(st.session_state.df1):,} 行</b>（主表）+ <b>{len(st.session_state.df2):,} 行</b>（数据源）
                    <br>已启用 <b>批量merge算法</b> 进行优化处理，预计耗时较短 🥔
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        
        with col_btn2:
            if st.button("🚀 开始比对与回填", type="primary", use_container_width=True):
                if not match_col1:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择主表匹配字段 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                if not match_col2:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择数据源匹配字段 🍠</div>
                    """, unsafe_allow_html=True)
                    return
                if not fill_cols:
                    st.markdown("""
                    <div class="error-cute">❌ 请至少选择一个回填字段 ✨</div>
                    """, unsafe_allow_html=True)
                    return
                
                # 创建进度显示区域
                progress_container = st.container()
                with progress_container:
                    progress_col1, progress_col2 = st.columns([3, 1])
                    with progress_col1:
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                    with progress_col2:
                        time_estimate = st.empty()
                    
                    info_text = st.empty()
                
                def update_progress(progress, stage=""):
                    progress_bar.progress(progress)
                    status_text.text(f"🥔 {stage} {int(progress * 100)}%")
                    # 简单的时间估算
                    if progress > 0:
                        elapsed = time.time() - update_progress.start_time if hasattr(update_progress, 'start_time') else 0
                        if elapsed > 0 and progress < 1:
                            estimated_total = elapsed / progress
                            remaining = estimated_total - elapsed
                            time_estimate.text(f"预计剩余: {int(remaining)}s")
                
                update_progress.start_time = time.time()
                
                with st.spinner("🍠 处理数据..."):
                    result_df, stats = compare_and_fill(
                        st.session_state.df1,
                        st.session_state.df2,
                        match_col1,
                        match_col2,
                        fill_cols,
                        update_progress
                    )
                
                # 清理进度显示
                progress_container.empty()
                
                st.session_state.result_df = result_df
                st.session_state.stats = stats
                
                # 统计结果
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="potato-card"><div class="potato-card-header">📊 处理结果统计</div></div>', unsafe_allow_html=True)
                
                result_col1, result_col2, result_col3, result_col4 = st.columns(4)
                
                with result_col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📝 总行数</div>
                        <div class="metric-value">{stats['total_rows']:,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with result_col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">✅ 匹配成功</div>
                        <div class="metric-value">{stats['matched_rows']:,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with result_col3:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">❌ 匹配失败</div>
                        <div class="metric-value">{stats['unmatched_rows']:,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with result_col4:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">✨ 回填单元格</div>
                        <div class="metric-value">{stats['filled_cells']:,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 匹配结果提示
                match_rate = (stats['matched_rows'] / stats['total_rows'] * 100) if stats['total_rows'] > 0 else 0
                
                if match_rate >= 80:
                    st.markdown(f"""
                    <div class="success-cute" style="font-size: 1rem; margin-top: 1rem;">
                        🎉 太棒了！匹配成功率 <strong>{match_rate:.1f}%</strong> 🥔🎉
                    </div>
                    """, unsafe_allow_html=True)
                elif match_rate >= 50:
                    st.markdown(f"""
                    <div class="warning-cute" style="font-size: 1rem; margin-top: 1rem;">
                        🤔 匹配成功率 <strong>{match_rate:.1f}%</strong>，部分土豆还没找到家 🍠
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div class="warning-cute" style="font-size: 1rem; margin-top: 1rem;">
                        😅 匹配成功率较低 (<strong>{match_rate:.1f}%</strong>)，请检查匹配字段配置 🥔
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown('<div class="potato-decoration" style="margin: 0.8rem 0;">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
                
                # 新增字段提示
                if stats.get('source_cols_added'):
                    added_cols_str = "、".join([f"`{col}`" for col in stats['source_cols_added']])
                    st.markdown(f"""
                    <div class="success-cute" style="margin-bottom: 0.5rem;">
                        📋 已新增数据源字段：{added_cols_str}
                    </div>
                    """, unsafe_allow_html=True)
                
                # 错误信息
                if stats['errors']:
                    with st.expander("🐛 查看提示信息"):
                        for error in stats['errors']:
                            st.markdown(f"<div class='warning-cute'>💡 {error}</div>", unsafe_allow_html=True)
                
                # 结果预览
                with st.expander("👁️ 预览处理结果"):
                    st.dataframe(result_df.head(50), use_container_width=True, height=380)
                
                # 下载按钮
                st.markdown("<hr>", unsafe_allow_html=True)
                
                excel_bytes = excel_to_bytes(result_df, "比对结果.xlsx")
                csv_bytes = csv_to_bytes(result_df, "比对结果.csv")
                
                # 添加导出格式选择
                export_format = st.radio(
                    "📥 选择导出格式",
                    options=["Excel (.xlsx)", "CSV (.csv)"],
                    horizontal=True,
                    help="选择下载文件的格式"
                )
                
                download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
                
                with download_col1:
                    st.markdown('<span style="font-size: 2rem;">🥔</span>', unsafe_allow_html=True)
                
                with download_col2:
                    if export_format == "Excel (.xlsx)":
                        st.download_button(
                            label="📥 下载结果Excel",
                            data=excel_bytes,
                            file_name="比对结果.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True
                        )
                    else:
                        st.download_button(
                            label="📥 下载结果CSV",
                            data=csv_bytes,
                            file_name="比对结果.csv",
                            mime="text/csv",
                            type="primary",
                            use_container_width=True
                        )
                
                with download_col3:
                    st.markdown('<span style="font-size: 2rem;">🍠</span>', unsafe_allow_html=True)
                
                st.markdown(f"""
                <div style="text-align: center; color: #8B4513; margin-top: 0.5rem;">
                    ⏱️ 处理耗时：<strong>{stats['processing_time']:.2f}</strong> 秒 | 
                    📊 文件：<strong>{len(result_df):,}</strong> 行 × <strong>{len(result_df.columns)}</strong> 列
                </div>
                """, unsafe_allow_html=True)
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


def compare_and_fill(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    match_col1: str,
    match_col2: str,
    fill_cols: list,
    progress_callback=None
) -> tuple[pd.DataFrame, dict]:
    """执行数据比对，使用 pandas merge 批量处理，结果包含主表所有字段 + 选中的回填字段（带_来源后缀）"""
    start_time = time.time()
    
    stats = {
        "total_rows": len(df1),
        "matched_rows": 0,
        "filled_cells": 0,
        "unmatched_rows": 0,
        "source_cols_added": [],
        "errors": []
    }
    
    if match_col1 not in df1.columns:
        stats["errors"].append(f"主表缺少匹配字段: {match_col1}")
        return df1.copy(), stats
    
    if match_col2 not in df2.columns:
        stats["errors"].append(f"数据源缺少匹配字段: {match_col2}")
        return df1.copy(), stats
    
    # 过滤掉无效的回填字段
    valid_fill_cols = [col for col in fill_cols if col in df2.columns]
    if len(valid_fill_cols) != len(fill_cols):
        invalid = [col for col in fill_cols if col not in df2.columns]
        stats["errors"].append(f"数据源缺少字段: {', '.join(invalid)}")
    
    if not valid_fill_cols:
        stats["errors"].append("没有有效的回填字段")
        return df1.copy(), stats
    
    # 更新进度：开始处理
    if progress_callback:
        progress_callback(0.2)
    
    # 准备数据源列（匹配列 + 回填列）
    source_cols = [match_col2] + valid_fill_cols
    df2_selected = df2[source_cols].copy()
    
    # 重命名数据源中的列，准备合并
    col_mapping = {match_col2: match_col1}  # 匹配列重命名
    renamed_fill_cols = []
    
    for col in valid_fill_cols:
        # 创建新列名（带_来源后缀）
        new_col = f"{col}_来源"
        # 确保新名称不与主表字段冲突
        counter = 1
        while new_col in df1.columns:
            new_col = f"{col}_来源{counter}"
            counter += 1
        
        col_mapping[col] = new_col
        renamed_fill_cols.append(new_col)
        stats["source_cols_added"].append(new_col)
    
    # 重命名数据源列
    df2_renamed = df2_selected.rename(columns=col_mapping)
    
    # 更新进度：准备合并
    if progress_callback:
        progress_callback(0.4)
    
    # 使用 pandas merge 批量处理
    # left join 确保主表所有行都保留
    result_df = df1.merge(
        df2_renamed,
        on=match_col1,
        how='left',
        suffixes=('', '')
    )
    
    # 更新进度：合并完成
    if progress_callback:
        progress_callback(0.7)
    
    # 统计信息
    # 匹配成功的行：在任意一个回填列中非空的行
    if renamed_fill_cols:
        # 检查是否有非空值
        is_matched = result_df[renamed_fill_cols].notna().any(axis=1)
        stats["matched_rows"] = int(is_matched.sum())
        stats["unmatched_rows"] = int((~is_matched).sum())
        
        # 统计回填的单元格数量
        stats["filled_cells"] = int(result_df[renamed_fill_cols].notna().sum().sum())
    else:
        stats["matched_rows"] = 0
        stats["unmatched_rows"] = len(result_df)
        stats["filled_cells"] = 0
    
    # 更新进度：完成
    if progress_callback:
        progress_callback(1.0)
    
    stats["processing_time"] = time.time() - start_time
    
    return result_df, stats


# ============================================
# 页面3：数据拆分器
# ============================================
def show_split_tool():
    """显示数据拆分器工具"""
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">✂️ 数据拆分器</h1>
        <p class="potato-subtitle">✨ 将大型Excel文件按指定条数拆分成多个文件 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">将大型Excel文件按指定条数拆分成多个小文件，适合数据分发和分批处理。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传Excel文件 → ② 设置拆分条数 → ③ 执行拆分 → ④ 下载压缩包
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 初始化session state
    if 'split_df' not in st.session_state:
        st.session_state.split_df = None
    if 'split_result' not in st.session_state:
        st.session_state.split_result = None
    
    # 使用说明
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0;">
            <span style="font-size: 2.5rem;">🥔</span>
            <h2 style="color: #8B4513; margin: 0.3rem 0;">使用说明</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="potato-card" style="margin-bottom: 0.8rem;">
            <div class="potato-card-header">🌱 操作步骤</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>上传 <b>Excel文件</b> 📁</li>
                <li>输入 <b>拆分条数</b> 🔢</li>
                <li>点击 <b>开始拆分</b> ✂️</li>
                <li>下载 <b>打包文件</b> 📥</li>
            </ol>
        </div>
        
        <div class="potato-card">
            <div class="potato-card-header">💡 温馨提示</div>
            <ul style="color: #8B4513; line-height: 1.7; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>拆分范围：10~100000条</li>
                <li>支持 .xlsx 和 .xls 格式</li>
                <li>结果打包成zip下载</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 2rem;">🥔 🌿</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("🥔 数据拆分器")
    
    # 文件上传区域
    st.markdown('<div class="potato-card"><div class="potato-card-header">📁 上传Excel文件</div></div>', unsafe_allow_html=True)
    
    file = st.file_uploader(
        "点击上传或拖拽Excel文件到此处",
        type=['xlsx', 'xls'],
        help="🥔 上传要拆分的Excel文件",
        key="split_file_uploader"
    )
    
    if file:
        with st.spinner("🥔 加载中..."):
            df = load_data_file(file)
            if df is not None:
                st.session_state.split_df = df
                st.session_state.split_result = None
                st.markdown("""
                <div class="success-cute">✅ 文件加载成功</div>
                """, unsafe_allow_html=True)
                
                # 显示文件信息
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="potato-card"><div class="potato-card-header">📊 文件信息</div></div>', unsafe_allow_html=True)
                
                info_col1, info_col2, info_col3 = st.columns(3)
                
                with info_col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📝 总行数</div>
                        <div class="metric-value">{len(df):,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📊 总列数</div>
                        <div class="metric-value">{len(df.columns)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col3:
                    file_size_mb = file.size / (1024 * 1024)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">💾 文件大小</div>
                        <div class="metric-value">{file_size_mb:.2f} MB</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 数据预览
                with st.expander("👁️ 预览数据（前20行）"):
                    st.dataframe(df.head(20), use_container_width=True, height=300)
    
    # 拆分配置
    if st.session_state.split_df is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 拆分配置</div></div>', unsafe_allow_html=True)
        
        config_col1, config_col2 = st.columns([1, 2])
        
        with config_col1:
            split_count = st.number_input(
                "📏 每份条数",
                min_value=10,
                max_value=100000,
                value=1000,
                step=100,
                help="每份文件的行数（10~100000）"
            )
        
        # 计算拆分结果
        total_rows = len(st.session_state.split_df)
        file_count = (total_rows + split_count - 1) // split_count
        
        with config_col2:
            st.markdown(f"""
            <div class="potato-card" style="margin-top: 0.5rem;">
                <div style="display: flex; justify-content: space-around; text-align: center;">
                    <div>
                        <div style="font-size: 0.9rem; color: #8B4513;">📝 原始数据</div>
                        <div style="font-size: 1.3rem; font-weight: 700; color: #D2691E;">{total_rows:,} 条</div>
                    </div>
                    <div style="color: #8B4513; font-size: 1.5rem;">→</div>
                    <div>
                        <div style="font-size: 0.9rem; color: #8B4513;">📁 拆分后</div>
                        <div style="font-size: 1.3rem; font-weight: 700; color: #D2691E;">{file_count} 个文件</div>
                    </div>
                    <div style="color: #8B4513; font-size: 1.5rem;">≈</div>
                    <div>
                        <div style="font-size: 0.9rem; color: #8B4513;">📏 每份约</div>
                        <div style="font-size: 1.3rem; font-weight: 700; color: #D2691E;">~{split_count:,} 条</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # 预估拆分详情
        st.markdown("<hr>", unsafe_allow_html=True)
        
        preview_col1, preview_col2 = st.columns(2)
        
        with preview_col1:
            st.markdown("**🥔 拆分预估**")
            st.markdown(f"""
            <div style="background: #FFF8DC; padding: 1rem; border-radius: 12px; color: #8B4513;">
                <p style="margin: 0.3rem 0;">• 预计生成 <strong>{file_count}</strong> 个Excel文件</p>
                <p style="margin: 0.3rem 0;">• 第1个文件：1 ~ {min(split_count, total_rows)} 条</p>
                <p style="margin: 0.3rem 0;">• 最后文件：{(file_count-1)*split_count + 1} ~ {total_rows} 条</p>
                <p style="margin: 0.3rem 0;">• 文件名格式：<code>data_0001.xlsx</code></p>
            </div>
            """, unsafe_allow_html=True)
        
        with preview_col2:
            st.markdown("**🍠 提示**")
            st.markdown("""
            <div style="background: #FFE4C4; padding: 1rem; border-radius: 12px; color: #8B4513;">
                <p style="margin: 0.3rem 0;">💡 数据将被保存为zip压缩包</p>
                <p style="margin: 0.3rem 0;">💡 每个Excel文件包含表头</p>
                <p style="margin: 0.3rem 0;">💡 建议拆分条数不要过大</p>
            </div>
            """, unsafe_allow_html=True)
        
        # 执行拆分按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        
        with col_btn2:
            if st.button("✂️ 开始拆分", type="primary", use_container_width=True):
                with st.spinner("🍠 正在拆分数据..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # 执行拆分
                    start_time = time.time()
                    split_dfs = []
                    total = len(st.session_state.split_df)
                    
                    for i in range(0, total, split_count):
                        chunk_df = st.session_state.split_df.iloc[i:i + split_count]
                        split_dfs.append(chunk_df)
                        
                        progress = min((i + split_count) / total, 1.0)
                        progress_bar.progress(progress)
                        status_text.text(f"🥔 拆分中... {int(progress * 100)}%")
                    
                    progress_bar.empty()
                    status_text.empty()
                    
                    # 生成zip文件
                    status_text = st.empty()
                    status_text.text("🍠 正在打包文件...")
                    
                    zip_bytes = excel_to_bytes_multi(split_dfs, "data")
                    
                    status_text.empty()
                    processing_time = time.time() - start_time
                    
                    st.session_state.split_result = {
                        'zip_bytes': zip_bytes,
                        'file_count': len(split_dfs),
                        'total_rows': total,
                        'split_count': split_count,
                        'processing_time': processing_time
                    }
                    
                    # 显示成功消息
                    st.markdown("""
                    <div class="success-cute" style="margin-top: 1rem;">
                        🎉 拆分完成！可以下载打包文件了 🥔🎉
                    </div>
                    """, unsafe_allow_html=True)
        
        # 显示下载区域
        if st.session_state.split_result:
            result = st.session_state.split_result
            
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown('<div class="potato-card"><div class="potato-card-header">📥 下载结果</div></div>', unsafe_allow_html=True)
            
            # 统计信息
            result_col1, result_col2, result_col3, result_col4 = st.columns(4)
            
            with result_col1:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-label">📁 文件数量</div>
                    <div class="metric-value">{result['file_count']}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with result_col2:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-label">📝 总行数</div>
                    <div class="metric-value">{result['total_rows']:,}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with result_col3:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-label">📏 每份条数</div>
                    <div class="metric-value">{result['split_count']:,}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with result_col4:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-label">⏱️ 处理耗时</div>
                    <div class="metric-value">{result['processing_time']:.2f}s</div>
                </div>
                """, unsafe_allow_html=True)
            
            # 下载按钮
            download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
            
            with download_col1:
                st.markdown('<span style="font-size: 2rem;">🥔</span>', unsafe_allow_html=True)
            
            with download_col2:
                st.download_button(
                    label="📥 下载ZIP打包文件",
                    data=result['zip_bytes'],
                    file_name=f"拆分数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip",
                    type="primary",
                    use_container_width=True
                )
            
            with download_col3:
                st.markdown('<span style="font-size: 2rem;">🍠</span>', unsafe_allow_html=True)
            
            # 文件大小提示
            zip_size_mb = len(result['zip_bytes']) / (1024 * 1024)
            st.markdown(f"""
            <div style="text-align: center; color: #8B4513; margin-top: 0.5rem;">
                📦 打包文件大小：<strong>{zip_size_mb:.2f}</strong> MB
            </div>
            """, unsafe_allow_html=True)
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 页面4：数据聚合器
# ============================================
def show_aggregate_tool():
    """显示数据聚合器工具"""
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🔗 数据聚合器</h1>
        <p class="potato-subtitle">✨ 将相同数据的行合并，让内容聚合更高效 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">将相同数据的行合并，让内容聚合更高效，适合数据汇总和去重场景。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传Excel文件 → ② 选择聚合字段 → ③ 选择待合并字段 → ④ 设置分隔符 → ⑤ 执行聚合
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 初始化session state
    if 'agg_df' not in st.session_state:
        st.session_state.agg_df = None
    if 'agg_result' not in st.session_state:
        st.session_state.agg_result = None
    
    # 使用说明
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0;">
            <span style="font-size: 2.5rem;">🥔</span>
            <h2 style="color: #8B4513; margin: 0.3rem 0;">使用说明</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="potato-card" style="margin-bottom: 0.8rem;">
            <div class="potato-card-header">🌱 操作步骤</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>上传 <b>Excel文件</b> 📁</li>
                <li>选择 <b>聚合字段</b> 🔑</li>
                <li>选择 <b>待合并字段</b> 📝</li>
                <li>设置 <b>分隔符</b> ✂️</li>
                <li>点击 <b>开始聚合</b> 🚀</li>
                <li>下载 <b>结果文件</b> 📥</li>
            </ol>
        </div>
        
        <div class="potato-card">
            <div class="potato-card-header">💡 示例说明</div>
            <ul style="color: #8B4513; line-height: 1.6; font-size: 0.85rem; padding-left: 1.2rem;">
                <li><b>聚合字段：</b>决定哪些行需要合并</li>
                <li><b>待合并字段：</b>内容会被连接起来</li>
                <li><b>分隔符：</b>连接时的分隔符号</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 2rem;">🥔 🌿</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("🥔 数据聚合器")
    
    # 文件上传区域
    st.markdown('<div class="potato-card"><div class="potato-card-header">📁 上传Excel文件</div></div>', unsafe_allow_html=True)
    
    file = st.file_uploader(
        "点击上传或拖拽Excel文件到此处",
        type=['xlsx', 'xls'],
        help="🥔 上传要聚合的Excel文件",
        key="agg_file_uploader"
    )
    
    if file:
        with st.spinner("🥔 加载中..."):
            df = load_data_file(file)
            if df is not None:
                st.session_state.agg_df = df
                st.session_state.agg_result = None
                st.markdown("""
                <div class="success-cute">✅ 文件加载成功</div>
                """, unsafe_allow_html=True)
                
                # 显示文件信息
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="potato-card"><div class="potato-card-header">📊 文件信息</div></div>', unsafe_allow_html=True)
                
                info_col1, info_col2, info_col3 = st.columns(3)
                
                with info_col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📝 总行数</div>
                        <div class="metric-value">{len(df):,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📊 总列数</div>
                        <div class="metric-value">{len(df.columns)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col3:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📋 字段列表</div>
                        <div class="metric-value">{len(df.columns)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 显示所有字段
                st.markdown("**📋 可用字段：**")
                fields_display = "、".join([f"`{col}`" for col in df.columns])
                st.markdown(f"<div style='color: #8B4513;'>{fields_display}</div>", unsafe_allow_html=True)
                
                # 数据预览
                with st.expander("👁️ 预览数据（前20行）"):
                    st.dataframe(df.head(20), use_container_width=True, height=300)
    
    # 聚合配置
    if st.session_state.agg_df is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 聚合配置</div></div>', unsafe_allow_html=True)
        
        # 字段选择
        config_col1, config_col2 = st.columns(2)
        
        with config_col1:
            agg_cols = st.multiselect(
                "🔑 聚合字段（相同值的行会合并）",
                options=list(st.session_state.agg_df.columns),
                default=[],
                help="选择用于分组的字段，这些字段完全相同的行会被合并"
            )
        
        with config_col2:
            merge_cols = st.multiselect(
                "📝 待合并字段（内容会拼接在一起）",
                options=[col for col in st.session_state.agg_df.columns if col not in agg_cols],
                default=[],
                help="选择要合并内容的字段"
            )
        
        # 分隔符设置
        sep_col1, sep_col2, sep_col3 = st.columns([1, 2, 1])
        
        with sep_col1:
            sep_options = {
                "逗号 ,": ",",
                "分号 ;": ";",
                "顿号 、": "、",
                "竖线 |": "|",
                "空格": " ",
                "换行（换行符）": "\n",
                "斜杠 /": "/",
                "自定义": "custom"
            }
            sep_preset = st.selectbox(
                "🔣 预设分隔符",
                options=list(sep_options.keys()),
                index=0
            )
        
        with sep_col2:
            if sep_preset == "自定义":
                separator = st.text_input(
                    "✏️ 自定义分隔符",
                    value=",",
                    max_chars=10,
                    help="输入自定义的分隔符"
                )
            else:
                separator = sep_options[sep_preset]
                st.text_input(
                    "🔣 分隔符预览",
                    value=f"「{separator}」",
                    disabled=True
                )
        
        with sep_col3:
            st.markdown("""
            <div style="padding: 0.8rem; background: #FFF8DC; border-radius: 12px; text-align: center;">
                <div style="color: #8B4513; font-size: 0.85rem;">💡 示例输出</div>
                <div style="color: #D2691E; font-weight: 600; margin-top: 0.3rem;">
                    A{sep}B{sep}C
                </div>
            </div>
            """.format(sep=separator if separator != "\n" else "换行"), unsafe_allow_html=True)
        
        # 配置验证提示
        if len(agg_cols) == 0 and len(merge_cols) == 0:
            st.markdown("""
            <div class="warning-cute" style="margin-top: 1rem;">
                🤔 请至少选择一个「聚合字段」和一个「待合并字段」来配置聚合规则 🥔
            </div>
            """, unsafe_allow_html=True)
        elif len(agg_cols) == 0:
            st.markdown("""
            <div class="warning-cute" style="margin-top: 1rem;">
                🤔 请至少选择一个「聚合字段」🥔
            </div>
            """, unsafe_allow_html=True)
        elif len(merge_cols) == 0:
            st.markdown("""
            <div class="warning-cute" style="margin-top: 1rem;">
                🤔 请至少选择一个「待合并字段」🍠
            </div>
            """, unsafe_allow_html=True)
        else:
            # 预估聚合结果
            st.markdown("<hr>", unsafe_allow_html=True)
            
            preview_col1, preview_col2 = st.columns(2)
            
            with preview_col1:
                # 计算预估结果
                original_count = len(st.session_state.agg_df)
                
                # 估算聚合后的行数（基于聚合字段的唯一组合数）
                if agg_cols:
                    unique_groups = st.session_state.agg_df.groupby(agg_cols).ngroups
                else:
                    unique_groups = 1
                
                st.markdown("**🥔 聚合预估**")
                st.markdown(f"""
                <div style="background: #FFF8DC; padding: 1rem; border-radius: 12px; color: #8B4513;">
                    <p style="margin: 0.3rem 0;">• 原始数据：<strong>{original_count:,}</strong> 条</p>
                    <p style="margin: 0.3rem 0;">• 聚合后预计：<strong>{unique_groups:,}</strong> 条</p>
                    <p style="margin: 0.3rem 0;">• 减少：<strong>{original_count - unique_groups:,}</strong> 条（{(1 - unique_groups/original_count)*100:.1f}%）</p>
                    <p style="margin: 0.3rem 0;">• 分隔符：<code>{separator if separator != chr(10) else '换行符'}</code></p>
                </div>
                """, unsafe_allow_html=True)
            
            with preview_col2:
                st.markdown("**🍠 提示**")
                st.markdown("""
                <div style="background: #FFE4C4; padding: 1rem; border-radius: 12px; color: #8B4513;">
                    <p style="margin: 0.3rem 0;">💡 聚合字段完全相同的行会被合并</p>
                    <p style="margin: 0.3rem 0;">💡 待合并字段的内容会按顺序拼接</p>
                    <p style="margin: 0.3rem 0;">💡 空值会自动跳过</p>
                    <p style="margin: 0.3rem 0;">💡 非待合并字段保留第一行值</p>
                </div>
                """, unsafe_allow_html=True)
        
        # 执行聚合按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        
        with col_btn2:
            if st.button("🚀 开始聚合", type="primary", use_container_width=True):
                if len(agg_cols) == 0:
                    st.markdown("""
                    <div class="error-cute">❌ 请至少选择一个「聚合字段」🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                if len(merge_cols) == 0:
                    st.markdown("""
                    <div class="error-cute">❌ 请至少选择一个「待合并字段」🍠</div>
                    """, unsafe_allow_html=True)
                    return
                
                with st.spinner("🍠 正在聚合数据..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    try:
                        # 执行聚合
                        start_time = time.time()
                        
                        status_text.text("🥔 正在分析数据结构...")
                        progress_bar.progress(0.2)
                        
                        df = st.session_state.agg_df.copy()
                        
                        # 确定需要保留的字段（非聚合且非待合并的字段）
                        other_cols = [col for col in df.columns if col not in agg_cols and col not in merge_cols]
                        
                        # 构建聚合函数
                        # - 聚合字段：保留（自然保留）
                        # - 待合并字段：用分隔符连接
                        # - 其他字段：取第一个值
                        agg_funcs = {}
                        for col in merge_cols:
                            agg_funcs[col] = lambda x, sep=separator: sep.join(
                                [str(v) for v in x.dropna().astype(str) if str(v).strip()]
                            )
                        for col in other_cols:
                            agg_funcs[col] = 'first'
                        
                        progress_bar.progress(0.4)
                        status_text.text("🥔 正在分组聚合...")
                        
                        # 执行groupby聚合
                        result_df = df.groupby(agg_cols, as_index=False, sort=False).agg(agg_funcs)
                        
                        progress_bar.progress(0.8)
                        status_text.text("🍠 正在整理结果...")
                        
                        # 调整列顺序：聚合字段 + 待合并字段 + 其他字段
                        final_columns = agg_cols + merge_cols + other_cols
                        result_df = result_df[[col for col in final_columns if col in result_df.columns]]
                        
                        progress_bar.progress(1.0)
                        status_text.empty()
                        progress_bar.empty()
                        
                        processing_time = time.time() - start_time
                        
                        st.session_state.agg_result = {
                            'result_df': result_df,
                            'original_count': len(st.session_state.agg_df),
                            'result_count': len(result_df),
                            'processing_time': processing_time
                        }
                        
                        # 显示成功消息
                        st.markdown("""
                        <div class="success-cute" style="margin-top: 1rem;">
                            🎉 聚合完成！可以下载结果文件了 🥔🎉
                        </div>
                        """, unsafe_allow_html=True)
                        
                    except Exception as e:
                        progress_bar.empty()
                        status_text.empty()
                        st.markdown(f"""
                        <div class="error-cute">
                            ❌ 聚合失败：{str(e)} 🥔
                        </div>
                        """, unsafe_allow_html=True)
    
    # 显示聚合结果
    if st.session_state.agg_result:
        result = st.session_state.agg_result
        result_df = result['result_df']
        
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📊 聚合结果统计</div></div>', unsafe_allow_html=True)
        
        # 统计卡片
        result_col1, result_col2, result_col3, result_col4, result_col5 = st.columns(5)
        
        with result_col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 原始行数</div>
                <div class="metric-value">{result['original_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 聚合后</div>
                <div class="metric-value">{result['result_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col3:
            reduced = result['original_count'] - result['result_count']
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📉 减少行数</div>
                <div class="metric-value">{reduced:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col4:
            reduce_rate = (1 - result['result_count'] / result['original_count']) * 100
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📊 压缩率</div>
                <div class="metric-value">{reduce_rate:.1f}%</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col5:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">⏱️ 处理时间</div>
                <div class="metric-value">{result['processing_time']:.2f}s</div>
            </div>
            """, unsafe_allow_html=True)
        
        # 聚合效果提示
        if reduce_rate > 50:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                🎉 太棒了！数据压缩了 <strong>{reduce_rate:.1f}%</strong>，效率大幅提升 🥔🎉
            </div>
            """, unsafe_allow_html=True)
        elif reduce_rate > 20:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                😊 不错的效果！数据压缩了 <strong>{reduce_rate:.1f}%</strong> 🍠
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="warning-cute" style="margin-top: 1rem;">
                🤔 数据相似度较低，压缩了 <strong>{reduce_rate:.1f}%</strong> 🥔
            </div>
            """, unsafe_allow_html=True)
        
        # 结果预览
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">👁️ 结果预览</div></div>', unsafe_allow_html=True)
        
        # 预览前50行
        preview_rows = min(50, len(result_df))
        st.dataframe(result_df.head(preview_rows), use_container_width=True, height=350)
        
        st.caption(f"显示前 {preview_rows} 行，共 {len(result_df):,} 行")
        
        # 下载按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📥 下载结果</div></div>', unsafe_allow_html=True)
        
        excel_bytes = excel_to_bytes(result_df, "聚合结果.xlsx")
        
        download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
        
        with download_col1:
            st.markdown('<span style="font-size: 2rem;">🥔</span>', unsafe_allow_html=True)
        
        with download_col2:
            st.download_button(
                label="📥 下载聚合结果Excel",
                data=excel_bytes,
                file_name=f"数据聚合结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        with download_col3:
            st.markdown('<span style="font-size: 2rem;">🍠</span>', unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="text-align: center; color: #8B4513; margin-top: 0.5rem;">
            📊 结果：<strong>{len(result_df):,}</strong> 行 × <strong>{len(result_df.columns)}</strong> 列
        </div>
        """, unsafe_allow_html=True)
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 页面5：域名提取器
# ============================================
def show_domain_tool():
    """显示域名提取器工具"""
    from urllib.parse import urlparse
    
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🌐 域名提取器</h1>
        <p class="potato-subtitle">✨ 从URL中提取主域名或子域名 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">从Excel中的URL列提取域名，支持政务类域名和普通域名，自动识别主域名或子域名。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传Excel文件 → ② 选择URL字段 → ③ 设置域名类型和提取类型 → ④ 执行提取 → ⑤ 下载结果
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 域名提取规则说明
    st.markdown("""
    <div class="potato-card" style="margin-bottom: 1rem;">
        <div class="potato-card-header">📖 提取规则说明</div>
        <div style="display: flex; flex-wrap: wrap; gap: 1rem; margin-top: 0.5rem;">
            <div style="flex: 1; min-width: 280px; background: #FFF8DC; padding: 0.8rem; border-radius: 10px;">
                <div style="font-weight: 700; color: #8B4513; margin-bottom: 0.5rem;">🏛️ 政务类域名（.gov.cn）</div>
                <table style="width: 100%; font-size: 0.85rem; color: #8B4513;">
                    <tr><td><b>示例URL：</b></td><td>https://services.credit.jiangsu.gov.cn:8809</td></tr>
                    <tr><td><b>主域名：</b></td><td><code style="background: #FFE4C4; padding: 0.1rem 0.3rem;">jiangsu.gov.cn</code>（最后三段）</td></tr>
                    <tr><td><b>子域名：</b></td><td><code style="background: #FFE4C4; padding: 0.1rem 0.3rem;">credit.jiangsu.gov.cn</code>（最后四段）</td></tr>
                </table>
            </div>
            <div style="flex: 1; min-width: 280px; background: #FFF0F5; padding: 0.8rem; border-radius: 10px;">
                <div style="font-weight: 700; color: #8B4513; margin-bottom: 0.5rem;">🌐 普通域名</div>
                <table style="width: 100%; font-size: 0.85rem; color: #8B4513;">
                    <tr><td><b>示例URL：</b></td><td>https://www.example.com/path/to/page</td></tr>
                    <tr><td><b>主域名：</b></td><td><code style="background: #FFE4C4; padding: 0.1rem 0.3rem;">example.com</code>（最后两段）</td></tr>
                    <tr><td><b>子域名：</b></td><td><code style="background: #FFE4C4; padding: 0.1rem 0.3rem;">www.example.com</code>（最后三段）</td></tr>
                </table>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 初始化session state
    if 'domain_df' not in st.session_state:
        st.session_state.domain_df = None
    if 'domain_result' not in st.session_state:
        st.session_state.domain_result = None
    
    # 域名提取函数 - 增加异常处理和中文错误提示
    def extract_domain_from_url(url):
        """从URL中提取域名（去除协议和端口号），返回 (域名, 错误信息) 元组"""
        try:
            # 处理空值
            if pd.isna(url) or not str(url).strip():
                return (None, "空值")
            
            url = str(url).strip()
            
            # 检查是否为空字符串
            if not url:
                return (None, "空值")
            
            # 检查是否包含有效域名特征（至少有一个点）
            if '.' not in url:
                return (None, "非URL格式")
            
            # 去除协议
            for protocol in ['https://', 'http://', 'HTTPS://', 'HTTP://']:
                if url.startswith(protocol):
                    url = url[len(protocol):]
                    break
            
            # 去除路径（取第一个/之前的部分）
            if '/' in url:
                url = url.split('/')[0]
            
            # 去除端口号
            if ':' in url:
                url = url.split(':')[0]
            
            # 验证结果是否有效域名
            if not url or len(url) < 4:  # 最少应该是 x.x 格式
                return (None, "域名格式无效")
            
            return (url, None)
            
        except Exception as e:
            return (None, f"解析异常")
    
    def is_gov_domain(domain):
        """判断是否为政务类域名"""
        try:
            if not domain:
                return False
            parts = domain.lower().split('.')
            # 政务类域名通常为 xxx.gov.cn 或 xxx.省.gov.cn 格式
            # 检查是否以 gov.cn 结尾
            if len(parts) >= 2 and parts[-2] == 'gov' and parts[-1] == 'cn':
                return True
            return False
        except Exception:
            return False
    
    def extract_target_domain(domain, domain_type, extract_type):
        """根据域名类型和提取类型提取目标域名，返回 (域名, 错误信息) 元组"""
        try:
            if not domain:
                return (None, "空域名")
            
            parts = domain.split('.')
            
            if domain_type == "政务类域名":
                # 政务类域名
                if extract_type == "主域名":
                    # 主域名 = 最后三段
                    if len(parts) >= 3:
                        return ('.'.join(parts[-3:]), None)
                    else:
                        return (domain, "域名段数不足(政务主域名)")
                else:
                    # 子域名 = 最后四段
                    if len(parts) >= 4:
                        return ('.'.join(parts[-4:]), None)
                    else:
                        return (domain, "域名段数不足(政务子域名)")
            else:
                # 普通域名
                if extract_type == "主域名":
                    # 主域名 = 最后两段
                    if len(parts) >= 2:
                        return ('.'.join(parts[-2:]), None)
                    else:
                        return (domain, "域名段数不足(普通主域名)")
                else:
                    # 子域名 = 最后三段
                    if len(parts) >= 3:
                        return ('.'.join(parts[-3:]), None)
                    else:
                        return (domain, "域名段数不足(普通子域名)")
                        
        except Exception as e:
            return (None, "提取异常")
    
    def safe_extract_domain(url, domain_type, extract_type):
        """安全的域名提取函数，统一处理所有异常"""
        try:
            # 第一步：提取域名（去除协议、端口、路径）
            domain, error1 = extract_domain_from_url(url)
            
            if error1:
                return (None, error1)
            
            # 第二步：根据配置提取目标域名
            result, error2 = extract_target_domain(domain, domain_type, extract_type)
            
            return (result, error2)
            
        except Exception as e:
            return (None, f"未知异常")
    
    # 使用说明
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0;">
            <span style="font-size: 2.5rem;">🥔</span>
            <h2 style="color: #8B4513; margin: 0.3rem 0;">使用说明</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="potato-card" style="margin-bottom: 0.8rem;">
            <div class="potato-card-header">🌱 操作步骤</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>上传 <b>Excel文件</b> 📁</li>
                <li>选择 <b>URL字段</b> 🔗</li>
                <li>设置 <b>域名类型</b> 🏛️</li>
                <li>设置 <b>提取类型</b> 🎯</li>
                <li>点击 <b>开始提取</b> 🚀</li>
                <li>下载 <b>结果文件</b> 📥</li>
            </ol>
        </div>
        
        <div class="potato-card">
            <div class="potato-card-header">💡 域名类型说明</div>
            <ul style="color: #8B4513; line-height: 1.6; font-size: 0.85rem; padding-left: 1.2rem;">
                <li><b>政务类域名：</b>.gov.cn结尾的政务网站</li>
                <li><b>普通域名：</b>商业/个人网站等</li>
            </ul>
        </div>
        
        <div class="potato-card" style="margin-top: 0.8rem;">
            <div class="potato-card-header">💡 提取类型说明</div>
            <ul style="color: #8B4513; line-height: 1.6; font-size: 0.85rem; padding-left: 1.2rem;">
                <li><b>主域名：</b>网站核心域名</li>
                <li><b>子域名：</b>包含部门/子站的域名</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 2rem;">🥔 🌿</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("🥔 域名提取器")
    
    # 文件上传区域
    st.markdown('<div class="potato-card"><div class="potato-card-header">📁 上传Excel文件</div></div>', unsafe_allow_html=True)
    
    file = st.file_uploader(
        "点击上传或拖拽Excel文件到此处",
        type=['xlsx', 'xls'],
        help="🥔 上传包含URL的Excel文件",
        key="domain_file_uploader"
    )
    
    if file:
        with st.spinner("🥔 加载中..."):
            df = load_data_file(file)
            if df is not None:
                st.session_state.domain_df = df
                st.session_state.domain_result = None
                st.markdown("""
                <div class="success-cute">✅ 文件加载成功</div>
                """, unsafe_allow_html=True)
                
                # 显示文件信息
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="potato-card"><div class="potato-card-header">📊 文件信息</div></div>', unsafe_allow_html=True)
                
                info_col1, info_col2, info_col3 = st.columns(3)
                
                with info_col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📝 总行数</div>
                        <div class="metric-value">{len(df):,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📊 总列数</div>
                        <div class="metric-value">{len(df.columns)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col3:
                    file_size_mb = file.size / (1024 * 1024)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">💾 文件大小</div>
                        <div class="metric-value">{file_size_mb:.2f} MB</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 显示所有字段
                st.markdown("**📋 可用字段：**")
                fields_display = "、".join([f"`{col}`" for col in df.columns])
                st.markdown(f"<div style='color: #8B4513;'>{fields_display}</div>", unsafe_allow_html=True)
                
                # 数据预览
                with st.expander("👁️ 预览数据（前20行）"):
                    st.dataframe(df.head(20), use_container_width=True, height=300)
    
    # 提取配置
    if st.session_state.domain_df is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 提取配置</div></div>', unsafe_allow_html=True)
        
        config_col1, config_col2, config_col3 = st.columns(3)
        
        with config_col1:
            url_field = st.selectbox(
                "🔗 选择URL字段",
                options=["（请选择）"] + list(st.session_state.domain_df.columns),
                index=0,
                help="选择包含URL的列"
            )
            if url_field == "（请选择）":
                url_field = None
        
        with config_col2:
            domain_type = st.radio(
                "🏛️ 域名类型",
                options=["政务类域名", "普通域名"],
                horizontal=True,
                help="选择URL所属的域名类型"
            )
        
        with config_col3:
            extract_type = st.radio(
                "🎯 提取类型",
                options=["主域名", "子域名"],
                horizontal=True,
                help="选择要提取的域名类型"
            )
        
        # 字段预览
        if url_field:
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown(f"**🥔 `{url_field}` 字段预览**")
            
            preview_df = st.session_state.domain_df[url_field].dropna().head(10)
            st.write(preview_df.tolist())
            
            # 统计空值
            null_count = st.session_state.domain_df[url_field].isnull().sum()
            st.caption(f"📊 共 {len(st.session_state.domain_df):,} 条记录，空值 {null_count:,} 条")
        
        # 执行提取按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        
        with col_btn2:
            if st.button("🚀 开始提取", type="primary", use_container_width=True):
                if not url_field:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择包含URL的字段 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                with st.spinner("🍠 正在提取域名..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    try:
                        start_time = time.time()
                        
                        status_text.text("🥔 正在解析URL...")
                        progress_bar.progress(0.2)
                        
                        df = st.session_state.domain_df.copy()
                        
                        # 确定新列名
                        if domain_type == "政务类域名":
                            if extract_type == "主域名":
                                new_col_name = "提取主域名(政务)"
                            else:
                                new_col_name = "提取子域名(政务)"
                        else:
                            if extract_type == "主域名":
                                new_col_name = "提取主域名"
                            else:
                                new_col_name = "提取子域名"
                        
                        status_text.text("🥔 正在提取域名...")
                        progress_bar.progress(0.4)
                        
                        # 使用安全的提取函数处理每一行
                        results = []
                        errors = []
                        total = len(df)
                        
                        for idx, url in enumerate(df[url_field]):
                            result, error = safe_extract_domain(url, domain_type, extract_type)
                            results.append(result)
                            errors.append(error)
                            
                            # 每100行更新一次进度
                            if idx % 100 == 0:
                                progress = 0.4 + (idx / total) * 0.4
                                progress_bar.progress(progress)
                        
                        progress_bar.progress(0.8)
                        status_text.text("🍠 正在整理结果...")
                        
                        # 添加结果列
                        df[new_col_name] = results
                        
                        # 统计成功/失败
                        success_count = sum(1 for e in errors if e is None)
                        fail_count = total - success_count
                        
                        # 统计错误类型
                        error_stats = {}
                        for e in errors:
                            if e is not None:
                                error_stats[e] = error_stats.get(e, 0) + 1
                        
                        progress_bar.progress(1.0)
                        status_text.empty()
                        progress_bar.empty()
                        
                        processing_time = time.time() - start_time
                        
                        st.session_state.domain_result = {
                            'result_df': df,
                            'success_count': success_count,
                            'fail_count': fail_count,
                            'total_count': total,
                            'new_col_name': new_col_name,
                            'processing_time': processing_time,
                            'domain_type': domain_type,
                            'extract_type': extract_type,
                            'error_stats': error_stats
                        }
                        
                        # 显示成功消息
                        st.markdown("""
                        <div class="success-cute" style="margin-top: 1rem;">
                            🎉 域名提取完成！可以下载结果文件了 🥔🎉
                        </div>
                        """, unsafe_allow_html=True)
                        
                    except Exception as e:
                        progress_bar.empty()
                        status_text.empty()
                        st.markdown(f"""
                        <div class="error-cute">
                            ❌ 提取失败：URL格式不正确，无法解析 🥔
                        </div>
                        """, unsafe_allow_html=True)
    
    # 显示提取结果
    if st.session_state.domain_result:
        result = st.session_state.domain_result
        result_df = result['result_df']
        
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📊 提取结果统计</div></div>', unsafe_allow_html=True)
        
        # 统计卡片
        result_col1, result_col2, result_col3, result_col4, result_col5 = st.columns(5)
        
        with result_col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 总行数</div>
                <div class="metric-value">{result['total_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">✅ 提取成功</div>
                <div class="metric-value" style="color: #228B22;">{result['success_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">⚠️ 跳过</div>
                <div class="metric-value" style="color: #FF6347;">{result['fail_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col4:
            success_rate = (result['success_count'] / result['total_count'] * 100) if result['total_count'] > 0 else 0
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📊 成功率</div>
                <div class="metric-value">{success_rate:.1f}%</div>
            </div>
            """, unsafe_allow_html=True)
        
        with result_col5:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">⏱️ 处理时间</div>
                <div class="metric-value">{result['processing_time']:.2f}s</div>
            </div>
            """, unsafe_allow_html=True)
        
        # 显示错误类型统计
        error_stats = result.get('error_stats', {})
        if error_stats:
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown('<div class="potato-card"><div class="potato-card-header">⚠️ 跳过详情（未能提取的记录）</div></div>', unsafe_allow_html=True)
            
            error_col1, error_col2 = st.columns([2, 1])
            
            with error_col1:
                # 显示错误统计表格
                error_data = []
                error_reason_map = {
                    "空值": "空单元格或空字符串",
                    "非URL格式": "不包含域名格式（如缺少'.'）",
                    "域名格式无效": "域名格式不符合规范",
                    "域名段数不足(政务主域名)": "政务域名段数不足（需要至少3段）",
                    "域名段数不足(政务子域名)": "政务域名段数不足（需要至少4段）",
                    "域名段数不足(普通主域名)": "普通域名段数不足（需要至少2段）",
                    "域名段数不足(普通子域名)": "普通域名段数不足（需要至少3段）",
                    "解析异常": "解析过程中发生异常",
                    "未知异常": "发生未知错误"
                }
                
                for error_type, count in sorted(error_stats.items(), key=lambda x: -x[1]):
                    reason = error_reason_map.get(error_type, error_type)
                    error_data.append({"跳过类型": error_type, "原因": reason, "数量": count})
                
                if error_data:
                    error_df = pd.DataFrame(error_data)
                    st.dataframe(error_df, use_container_width=True, hide_index=True)
            
            with error_col2:
                st.markdown("""
                <div style="background: #FFF8DC; padding: 1rem; border-radius: 12px;">
                    <div style="color: #8B4513; font-size: 0.85rem; font-weight: 600; margin-bottom: 0.5rem;">💡 说明</div>
                    <ul style="color: #8B4513; font-size: 0.8rem; line-height: 1.6; padding-left: 1.2rem; margin: 0;">
                        <li>空值行已自动跳过</li>
                        <li>异常值已自动跳过</li>
                        <li>结果中显示为空</li>
                        <li>不会中断处理流程</li>
                    </ul>
                </div>
                """, unsafe_allow_html=True)
        
        # 提取效果提示
        if success_rate >= 95:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                🎉 太棒了！提取成功率 <strong>{success_rate:.1f}%</strong> 🥔🎉
            </div>
            """, unsafe_allow_html=True)
        elif success_rate >= 80:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                😊 不错的效果！提取成功率 <strong>{success_rate:.1f}%</strong> 🍠
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="warning-cute" style="margin-top: 1rem;">
                🤔 提取成功率 <strong>{success_rate:.1f}%</strong>，请检查URL格式是否正确 🥔
            </div>
            """, unsafe_allow_html=True)
        
        # 显示提取配置
        st.markdown(f"""
        <div style="background: #FFF8DC; padding: 0.8rem; border-radius: 12px; margin-top: 0.5rem;">
            <div style="color: #8B4513; font-size: 0.9rem;">
                <b>⚙️ 提取配置：</b>
                域名类型：<code>{result['domain_type']}</code> | 
                提取类型：<code>{result['extract_type']}</code> | 
                新增列：<code>{result['new_col_name']}</code>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 结果预览
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">👁️ 结果预览</div></div>', unsafe_allow_html=True)
        
        # 显示新增列
        preview_cols = [col for col in result_df.columns if col != result['new_col_name']][:3] + [result['new_col_name']]
        preview_cols = [col for col in preview_cols if col in result_df.columns]
        
        # 预览前50行
        preview_rows = min(50, len(result_df))
        st.dataframe(result_df.head(preview_rows), use_container_width=True, height=350)
        
        st.caption(f"显示前 {preview_rows} 行，共 {len(result_df):,} 行 | 结果列：<code>{result['new_col_name']}</code>")
        
        # 下载按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📥 下载结果</div></div>', unsafe_allow_html=True)
        
        excel_bytes = excel_to_bytes(result_df, "域名提取结果.xlsx")
        
        download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
        
        with download_col1:
            st.markdown('<span style="font-size: 2rem;">🥔</span>', unsafe_allow_html=True)
        
        with download_col2:
            st.download_button(
                label="📥 下载提取结果Excel",
                data=excel_bytes,
                file_name=f"域名提取结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        with download_col3:
            st.markdown('<span style="font-size: 2rem;">🍠</span>', unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="text-align: center; color: #8B4513; margin-top: 0.5rem;">
            📊 结果：<strong>{len(result_df):,}</strong> 行 × <strong>{len(result_df.columns)}</strong> 列
        </div>
        """, unsafe_allow_html=True)
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 页面6：单位树构建器
# ============================================

# 定义一级分组常量
GROUP_LIST = ["[党委]", "[政府]", "[人大]", "[政协]", "[法院]", "[检察院]", "[群众团体]", "[国有企业]", "[直联企业]", "[民营企业]", "[待分组]"]


def classify_group(unit_name: str, unit_type: str) -> str:
    """根据单位性质和名称关键词判断分组
    
    Args:
        unit_name: 单位名称
        unit_type: 单位性质（国有企业/企业/党政机关/群团等）
    
    Returns:
        分组名称，如 [政府]、[党委] 等
    """
    if pd.isna(unit_name):
        unit_name = ""
    if pd.isna(unit_type):
        unit_type = ""
    
    unit_name = str(unit_name).strip()
    unit_type = str(unit_type).strip()
    
    # 1. 国有企业 -> [国有企业]
    if unit_type == "国有企业":
        return "[国有企业]"
    
    # 2. 企业（非国企）-> [直联企业]
    if unit_type == "企业":
        return "[直联企业]"
    
    # 3. 群团 -> [群众团体]
    if unit_type == "群团":
        return "[群众团体]"
    
    # 4. 党政机关 -> 根据关键词判断
    if unit_type == "党政机关":
        # 优先检查关键词顺序很重要
        # 检察院
        if "检察院" in unit_name:
            return "[检察院]"
        # 法院
        if "法院" in unit_name:
            return "[法院]"
        # 政协
        if "政协" in unit_name:
            return "[政协]"
        # 人大
        if "人大" in unit_name:
            return "[人大]"
        # 纪委（包含纪委的组织）
        if "纪委" in unit_name or "纪检" in unit_name:
            return "[党委]"
        # 组织部、宣传部等党委部门
        if any(keyword in unit_name for keyword in ["党委", "组织部", "宣传部", "统战部", "政法委", "编办", "直属机关", "党校"]):
            return "[党委]"
        # 政府关键词
        if any(keyword in unit_name for keyword in ["厅", "局", "委", "办", "政府"]):
            return "[政府]"
    
    # 默认归入[待分组]
    return "[待分组]"


def is_valid_parent(parent: str, all_units: set) -> bool:
    """判断上级节点是否有效
    
    Args:
        parent: 上级节点名称
        all_units: 所有单位名称集合
    
    Returns:
        是否有效
    """
    if pd.isna(parent) or not str(parent).strip():
        return False
    
    parent = str(parent).strip()
    
    # 空字符串无效
    if not parent:
        return False
    
    # 如果是分组名称（如[政府]），有效
    if parent in GROUP_LIST:
        return True
    
    # 如果是某个单位名称，有效
    if parent in all_units:
        return True
    
    return False


def get_parent_node(unit_name: str, admin_unit: str, all_units: set, group: str) -> str:
    """判断上级节点
    
    Args:
        unit_name: 单位名称
        admin_unit: 行政主管单位
        all_units: 所有单位名称集合
        group: 该单位的分组
    
    Returns:
        上级节点名称
    """
    # 情况1：行政主管单位 = 单位名称本身（根节点）
    if not pd.isna(admin_unit) and str(admin_unit).strip() == str(unit_name).strip():
        return group
    
    # 情况2：行政主管单位在单位列表中存在
    if not pd.isna(admin_unit) and str(admin_unit).strip() in all_units:
        return str(admin_unit).strip()
    
    # 情况3：行政主管单位是一级分组名称
    if not pd.isna(admin_unit) and str(admin_unit).strip() in GROUP_LIST:
        return str(admin_unit).strip()
    
    # 无效情况
    return "[待分组]"


def calculate_level(unit_name: str, parent_node: str, all_units: set, level_cache: dict) -> int:
    """计算单位层级（递归）
    
    Args:
        unit_name: 单位名称
        parent_node: 上级节点
        all_units: 所有单位名称集合
        level_cache: 层级缓存
    
    Returns:
        层级（根节点为1级）
    """
    if unit_name in level_cache:
        return level_cache[unit_name]
    
    # 如果上级是分组，根节点为1级
    if parent_node in GROUP_LIST:
        level_cache[unit_name] = 1
        return 1
    
    # 如果上级是其他单位，递归计算
    if parent_node in all_units:
        # 需要找到 parent_node 的上级
        # 这里简化处理，假设最多几层
        for _, row in pd.DataFrame({"name": [unit_name], "parent": [parent_node]}).iter():
            parent_parent = parent_node
            level = 2  # 有上级单位，至少是2级
            # 简单处理：找到parent的上级
            if parent_parent in all_units and parent_parent not in GROUP_LIST:
                level += 1
            level_cache[unit_name] = level
            return level
    
    # 默认情况
    level_cache[unit_name] = 1
    return 1


def show_unit_tree_tool():
    """显示单位树构建器工具"""
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🌳 单位树构建器</h1>
        <p class="potato-subtitle">✨ 根据单位数据自动构建组织架构树 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">根据单位名称、行政主管单位、单位性质等字段，自动构建组织架构树，确定上级节点和分组归属。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传Excel/CSV文件 → ② 确认字段映射 → ③ 点击构建 → ④ 查看统计与预览 → ⑤ 下载结果
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 分组规则说明
    with st.expander("📖 分组与上级节点规则说明", expanded=False):
        st.markdown("""
        <div class="potato-card">
            <div class="potato-card-header">🌳 一级分组（共10个）</div>
            <div style="display: grid; grid-template-columns: repeat(5, 1fr); gap: 0.5rem; margin-top: 0.5rem;">
                <div style="background: #FFF8DC; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[党委]</div>
                <div style="background: #E8F5E9; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[政府]</div>
                <div style="background: #F3E5F5; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[人大]</div>
                <div style="background: #E3F2FD; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[政协]</div>
                <div style="background: #FFEBEE; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[法院]</div>
                <div style="background: #FFF3E0; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[检察院]</div>
                <div style="background: #F1F8E9; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[群众团体]</div>
                <div style="background: #FCE4EC; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[国有企业]</div>
                <div style="background: #E0F2F1; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[直联企业]</div>
                <div style="background: #ECEFF1; padding: 0.5rem; border-radius: 8px; text-align: center; color: #8B4513;">[待分组]</div>
            </div>
        </div>
        
        <div style="margin-top: 1rem;" class="potato-card">
            <div class="potato-card-header">📋 分组判定规则</div>
            <table style="width: 100%; font-size: 0.85rem; color: #8B4513;">
                <tr style="background: #FFF8DC;"><td><b>单位性质</b></td><td><b>名称关键词</b></td><td><b>分组</b></td></tr>
                <tr><td>国有企业</td><td>任意</td><td>[国有企业]</td></tr>
                <tr><td>企业</td><td>任意</td><td>[直联企业]</td></tr>
                <tr><td>党政机关</td><td>法院</td><td>[法院]</td></tr>
                <tr><td>党政机关</td><td>检察院</td><td>[检察院]</td></tr>
                <tr><td>党政机关</td><td>政协</td><td>[政协]</td></tr>
                <tr><td>党政机关</td><td>人大</td><td>[人大]</td></tr>
                <tr><td>党政机关</td><td>党委、纪委、组织部、宣传部等</td><td>[党委]</td></tr>
                <tr><td>党政机关</td><td>厅、局、委、办、政府</td><td>[政府]</td></tr>
                <tr><td>群团</td><td>任意</td><td>[群众团体]</td></tr>
            </table>
        </div>
        
        <div style="margin-top: 1rem;" class="potato-card">
            <div class="potato-card-header">🔗 上级节点判定规则</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.5rem;">
                <li><b>行政主管单位 = 单位名称本身</b> → 上级节点 = 该单位所属分组（作为根节点）</li>
                <li><b>行政主管单位在单位列表中存在</b> → 上级节点 = 行政主管单位</li>
                <li><b>行政主管单位是一级分组名称</b> → 上级节点 = 该分组</li>
                <li><b>其他情况</b> → 上级节点 = [待分组]</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)
    
    # 初始化session state
    if 'tree_df' not in st.session_state:
        st.session_state.tree_df = None
    if 'tree_result' not in st.session_state:
        st.session_state.tree_result = None
    
    # 使用说明
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0;">
            <span style="font-size: 2.5rem;">🥔</span>
            <h2 style="color: #8B4513; margin: 0.3rem 0;">使用说明</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="potato-card" style="margin-bottom: 0.8rem;">
            <div class="potato-card-header">🌱 操作步骤</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>上传 <b>Excel/CSV文件</b> 📁</li>
                <li>确认 <b>字段映射</b> 🔍</li>
                <li>点击 <b>开始构建</b> 🌳</li>
                <li>查看 <b>统计与预览</b> 📊</li>
                <li>下载 <b>结果文件</b> 📥</li>
            </ol>
        </div>
        
        <div class="potato-card">
            <div class="potato-card-header">💡 字段要求</div>
            <ul style="color: #8B4513; line-height: 1.7; font-size: 0.85rem; padding-left: 1.2rem;">
                <li><b>单位名称：</b>必填，单位全称</li>
                <li><b>行政主管单位：</b>可为空</li>
                <li><b>单位性质：</b>如"党政机关"</li>
                <li><b>区域：</b>如"XX区"</li>
            </ul>
        </div>
        
        <div class="potato-card" style="margin-top: 0.8rem;">
            <div class="potato-card-header">💡 支持格式</div>
            <ul style="color: #8B4513; line-height: 1.7; font-size: 0.85rem; padding-left: 1.2rem;">
                <li>.xlsx / .xls / .csv</li>
                <li>CSV自动检测编码</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 2rem;">🥔 🌿</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("🥔 单位树构建器")
    
    # 文件上传区域
    st.markdown('<div class="potato-card"><div class="potato-card-header">📁 上传单位数据文件</div></div>', unsafe_allow_html=True)
    
    file = st.file_uploader(
        "点击上传或拖拽Excel/CSV文件到此处",
        type=['xlsx', 'xls', 'csv'],
        help="🥔 上传包含单位数据的文件",
        key="tree_file_uploader"
    )
    
    if file:
        with st.spinner("🥔 加载中..."):
            df = load_data_file(file)
            if df is not None:
                st.session_state.tree_df = df
                st.session_state.tree_result = None
                st.markdown("""
                <div class="success-cute">✅ 文件加载成功</div>
                """, unsafe_allow_html=True)
                
                # 显示文件信息
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="potato-card"><div class="potato-card-header">📊 文件信息</div></div>', unsafe_allow_html=True)
                
                info_col1, info_col2, info_col3 = st.columns(3)
                
                with info_col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📝 总行数</div>
                        <div class="metric-value">{len(df):,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📊 总列数</div>
                        <div class="metric-value">{len(df.columns)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col3:
                    file_size_mb = file.size / (1024 * 1024)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">💾 文件大小</div>
                        <div class="metric-value">{file_size_mb:.2f} MB</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 显示所有字段
                st.markdown("**📋 可用字段：**")
                fields_display = "、".join([f"`{col}`" for col in df.columns])
                st.markdown(f"<div style='color: #8B4513;'>{fields_display}</div>", unsafe_allow_html=True)
                
                # 数据预览
                with st.expander("👁️ 预览数据（前20行）"):
                    st.dataframe(df.head(20), use_container_width=True, height=300)
    
    # 字段映射配置
    if st.session_state.tree_df is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 字段映射配置</div></div>', unsafe_allow_html=True)
        
        st.markdown("""
        <div style="background: #FFF8DC; padding: 0.8rem; border-radius: 10px; margin-bottom: 1rem;">
            <div style="color: #8B4513; font-size: 0.9rem;">
                💡 请为每个必填字段选择对应的列。系统会自动识别相似名称的字段。
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 自动识别字段
        cols = list(st.session_state.tree_df.columns)
        
        # 单位名称字段
        default_unit = None
        for col in cols:
            if any(keyword in col for keyword in ["单位名称", "单位", "名称", "name"]):
                default_unit = col
                break
        
        # 行政主管单位字段
        default_admin = None
        for col in cols:
            if any(keyword in col for keyword in ["行政主管", "主管单位", "上级", "parent"]):
                default_admin = col
                break
        
        # 单位性质字段
        default_type = None
        for col in cols:
            if any(keyword in col for keyword in ["单位性质", "性质", "type"]):
                default_type = col
                break
        
        # 区域字段
        default_area = None
        for col in cols:
            if any(keyword in col for keyword in ["区域", "区县", "area", "区域名称"]):
                default_area = col
                break
        
        config_col1, config_col2 = st.columns(2)
        
        with config_col1:
            unit_name_col = st.selectbox(
                "🏷️ 单位名称字段（必填）",
                options=["（请选择）"] + cols,
                index=(cols.index(default_unit) + 1) if default_unit and default_unit in cols else 0,
                help="选择包含单位名称的列"
            )
            if unit_name_col == "（请选择）":
                unit_name_col = None
            
            unit_type_col = st.selectbox(
                "📋 单位性质字段",
                options=["（不映射）"] + cols,
                index=(cols.index(default_type) + 1) if default_type and default_type in cols else 0,
                help="选择包含单位性质的列（如：党政机关、企业、国有企业等）"
            )
            if unit_type_col == "（不映射）":
                unit_type_col = None
        
        with config_col2:
            admin_unit_col = st.selectbox(
                "🔗 行政主管单位字段",
                options=["（不映射）"] + cols,
                index=(cols.index(default_admin) + 1) if default_admin and default_admin in cols else 0,
                help="选择包含行政主管单位的列"
            )
            if admin_unit_col == "（不映射）":
                admin_unit_col = None
            
            area_col = st.selectbox(
                "🗺️ 区域字段",
                options=["（不映射）"] + cols,
                index=(cols.index(default_area) + 1) if default_area and default_area in cols else 0,
                help="选择包含区域/区县名称的列"
            )
            if area_col == "（不映射）":
                area_col = None
        
        # 验证必填字段
        if unit_name_col is None:
            st.markdown("""
            <div class="warning-cute" style="margin-top: 1rem;">
                ⚠️ 请至少选择「单位名称」字段作为必填项 🥔
            </div>
            """, unsafe_allow_html=True)
        
        # 字段预览
        if unit_name_col:
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown(f"**🥔 单位名称字段预览**")
            
            preview_df = st.session_state.tree_df[unit_name_col].dropna().head(10)
            st.write(preview_df.tolist())
            
            # 统计空值
            null_count = st.session_state.tree_df[unit_name_col].isnull().sum()
            total_count = len(st.session_state.tree_df)
            st.caption(f"📊 共 {total_count:,} 条记录，空值 {null_count:,} 条")
        
        # 执行构建按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        
        with col_btn2:
            if st.button("🌳 开始构建单位树", type="primary", use_container_width=True):
                if unit_name_col is None:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择「单位名称」字段 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                with st.spinner("🍠 正在构建单位树..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    try:
                        start_time = time.time()
                        
                        status_text.text("🥔 正在处理数据...")
                        progress_bar.progress(0.1)
                        
                        df = st.session_state.tree_df.reset_index(drop=True).copy()
                        
                        # 验证列名是否存在
                        if unit_name_col not in df.columns:
                            progress_bar.empty()
                            status_text.empty()
                            st.markdown(f"""
                            <div class="error-cute">❌ 列 '{unit_name_col}' 不存在，请重新选择 🥔</div>
                            """, unsafe_allow_html=True)
                            return
                        
                        # 过滤掉单位名称为空的行
                        original_count = len(df)
                        df = df[df[unit_name_col].notna() & (df[unit_name_col].astype(str).str.strip() != "")]
                        
                        # 去重：同一单位名称只保留一条
                        df = df.drop_duplicates(subset=[unit_name_col], keep='first')
                        df = df.reset_index(drop=True)  # 重置索引
                        after_dedup = len(df)
                        
                        status_text.text("🥔 正在分析数据结构...")
                        progress_bar.progress(0.2)
                        
                        # 构建所有单位名称集合
                        all_units = set(df[unit_name_col].astype(str).str.strip())
                        
                        # 获取单位性质（用于分组判定）
                        if unit_type_col and unit_type_col in df.columns:
                            df['单位性质_处理'] = df[unit_type_col].fillna('')
                        else:
                            df['单位性质_处理'] = ''
                        
                        progress_bar.progress(0.3)
                        status_text.text("🥔 正在分组...")
                        
                        # 分组判定
                        groups = []
                        for idx in range(len(df)):
                            unit_name = df.loc[idx, unit_name_col]
                            unit_type = df.loc[idx, '单位性质_处理']
                            group = classify_group(unit_name, unit_type)
                            groups.append(group)
                        
                        df['分组'] = groups
                        
                        progress_bar.progress(0.5)
                        status_text.text("🍠 正在确定上级节点...")
                        
                        # 上级节点判定
                        parent_nodes = []
                        for idx in range(len(df)):
                            unit_name = df.loc[idx, unit_name_col]
                            admin_unit = df.loc[idx, admin_unit_col] if admin_unit_col and admin_unit_col in df.columns else None
                            group = df.loc[idx, '分组']
                            
                            parent = get_parent_node(unit_name, admin_unit, all_units, group)
                            parent_nodes.append(parent)
                        
                        df['上级节点'] = parent_nodes
                        
                        progress_bar.progress(0.7)
                        status_text.text("🍠 正在计算层级...")
                        
                        # 简化层级计算
                        levels = []
                        for idx in range(len(df)):
                            parent = df.loc[idx, '上级节点']
                            unit_name = df.loc[idx, unit_name_col]
                            
                            if parent in GROUP_LIST:
                                levels.append(1)
                            elif parent in all_units and parent != unit_name:
                                levels.append(2)  # 简化处理
                            else:
                                levels.append(1)
                        
                        df['层级'] = levels
                        
                        progress_bar.progress(0.85)
                        status_text.text("🍠 正在整理结果...")
                        
                        # 添加区域字段
                        if area_col and area_col in df.columns:
                            df['区域'] = df[area_col].fillna('未知')
                        else:
                            df['区域'] = '未知'
                        
                        progress_bar.progress(0.95)
                        status_text.text("🍠 正在生成统计...")
                        
                        # 构建最终结果
                        result_columns = [unit_name_col, '上级节点', '层级', '分组', '区域']
                        if admin_unit_col and admin_unit_col in df.columns:
                            result_columns.insert(1, admin_unit_col)
                        if unit_type_col and unit_type_col in df.columns:
                            result_columns.append(unit_type_col)
                        
                        # 确保所有列都存在
                        available_cols = [col for col in result_columns if col in df.columns or col in ['上级节点', '层级', '分组', '区域']]
                        result_df = df[available_cols].copy()
                        
                        # 重命名列
                        column_rename = {
                            unit_name_col: '单位名称',
                            admin_unit_col: '行政主管单位' if admin_unit_col else None,
                            unit_type_col: '单位性质' if unit_type_col else None
                        }
                        column_rename = {k: v for k, v in column_rename.items() if v is not None}
                        result_df = result_df.rename(columns=column_rename)
                        
                        # 统计信息
                        group_stats = result_df['分组'].value_counts().to_dict()
                        area_stats = result_df['区域'].value_counts().to_dict() if '区域' in result_df.columns else {}
                        
                        # 待分组统计
                        pending_count = len(result_df[result_df['上级节点'] == '[待分组]'])
                        
                        progress_bar.progress(1.0)
                        status_text.empty()
                        progress_bar.empty()
                        
                        processing_time = time.time() - start_time
                        
                        st.session_state.tree_result = {
                            'result_df': result_df,
                            'original_count': original_count,
                            'after_dedup': after_dedup,
                            'group_stats': group_stats,
                            'area_stats': area_stats,
                            'pending_count': pending_count,
                            'processing_time': processing_time,
                            'unit_name_col': '单位名称',
                            'admin_unit_col': '行政主管单位' if admin_unit_col else None,
                            'unit_type_col': '单位性质' if unit_type_col else None
                        }
                        
                        # 显示成功消息
                        st.markdown("""
                        <div class="success-cute" style="margin-top: 1rem;">
                            🎉 单位树构建完成！可以下载结果文件了 🥔🎉
                        </div>
                        """, unsafe_allow_html=True)
                        
                    except Exception as e:
                        progress_bar.empty()
                        status_text.empty()
                        st.markdown(f"""
                        <div class="error-cute">
                            ❌ 构建失败：{str(e)} 🥔
                        </div>
                        """, unsafe_allow_html=True)
    
    # 显示构建结果
    if st.session_state.tree_result:
        result = st.session_state.tree_result
        result_df = result['result_df']
        
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📊 构建结果统计</div></div>', unsafe_allow_html=True)
        
        # 统计卡片
        stat_col1, stat_col2, stat_col3, stat_col4, stat_col5 = st.columns(5)
        
        with stat_col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 原始行数</div>
                <div class="metric-value">{result['original_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 去重后</div>
                <div class="metric-value">{result['after_dedup']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">🏷️ 分组数量</div>
                <div class="metric-value">{len(result['group_stats']):,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col4:
            pending = result['pending_count']
            color = "#FF6347" if pending > 0 else "#228B22"
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">⚠️ 待分组</div>
                <div class="metric-value" style="color: {color};">{pending:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col5:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">⏱️ 处理时间</div>
                <div class="metric-value">{result['processing_time']:.2f}s</div>
            </div>
            """, unsafe_allow_html=True)
        
        # 分组统计详情
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📋 分组统计</div></div>', unsafe_allow_html=True)
        
        # 分组统计表格
        group_data = []
        for group in GROUP_LIST:
            if group in result['group_stats']:
                group_data.append({
                    "分组": group,
                    "单位数量": result['group_stats'][group]
                })
        
        if group_data:
            group_df = pd.DataFrame(group_data)
            group_df = group_df.sort_values('单位数量', ascending=False)
            
            # 显示分组统计
            display_cols = st.columns(5)
            for i, (_, row) in enumerate(group_df.iterrows()):
                with display_cols[i % 5]:
                    st.markdown(f"""
                    <div class="metric-card" style="padding: 0.6rem;">
                        <div class="metric-label" style="font-size: 0.75rem;">{row['分组']}</div>
                        <div class="metric-value">{row['单位数量']}</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            with st.expander("📊 查看分组统计表格"):
                st.dataframe(group_df, use_container_width=True, hide_index=True)
        
        # 区域统计（如果存在）
        if result['area_stats'] and len(result['area_stats']) > 1:
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown('<div class="potato-card"><div class="potato-card-header">🗺️ 区域统计</div></div>', unsafe_allow_html=True)
            
            area_data = []
            for area, count in sorted(result['area_stats'].items(), key=lambda x: -x[1]):
                area_data.append({
                    "区域": area,
                    "单位数量": count
                })
            
            if area_data:
                with st.expander("📊 查看区域统计"):
                    area_df = pd.DataFrame(area_data)
                    st.dataframe(area_df, use_container_width=True, hide_index=True)
        
        # 待分组详情
        if result['pending_count'] > 0:
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown('<div class="potato-card"><div class="potato-card-header">⚠️ 待分组单位详情</div></div>', unsafe_allow_html=True)
            
            pending_df = result_df[result_df['上级节点'] == '[待分组]'][['单位名称']].head(20)
            
            with st.expander(f"📋 查看待分组单位（共 {result['pending_count']} 个，显示前20个）"):
                st.dataframe(pending_df, use_container_width=True, hide_index=True)
        
        # 效果提示
        pending_rate = (result['pending_count'] / result['after_dedup'] * 100) if result['after_dedup'] > 0 else 0
        
        if pending_rate <= 5:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                🎉 太棒了！97%+ 的单位已成功分组 🥔🎉
            </div>
            """, unsafe_allow_html=True)
        elif pending_rate <= 20:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                😊 不错的效果！{100-pending_rate:.1f}% 的单位已成功分组 🍠
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="warning-cute" style="margin-top: 1rem;">
                🤔 有 {pending_rate:.1f}% 的单位需要检查行政主管单位配置 🥔
            </div>
            """, unsafe_allow_html=True)
        
        # 结果预览
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">👁️ 结果预览</div></div>', unsafe_allow_html=True)
        
        # 预览前50行
        preview_rows = min(50, len(result_df))
        
        # 添加筛选功能
        filter_col1, filter_col2 = st.columns(2)
        
        with filter_col1:
            filter_group = st.multiselect(
                "🔍 按分组筛选",
                options=GROUP_LIST,
                default=[],
                help="筛选特定分组的单位"
            )
        
        with filter_col2:
            filter_pending = st.checkbox("⚠️ 只显示待分组", value=False, help="只显示待分组的单位")
        
        # 应用筛选
        display_df = result_df.copy()
        if filter_group:
            display_df = display_df[display_df['分组'].isin(filter_group)]
        if filter_pending:
            display_df = display_df[display_df['上级节点'] == '[待分组]']
        
        st.markdown(f"📊 共 **{len(display_df):,}** 条记录（原始 **{len(result_df):,}** 条）")
        
        st.dataframe(display_df.head(preview_rows), use_container_width=True, height=350)
        
        st.caption(f"显示前 {min(preview_rows, len(display_df))} 行")
        
        # 下载按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📥 下载结果</div></div>', unsafe_allow_html=True)
        
        # 导出格式选择
        export_format = st.radio(
            "📥 选择导出格式",
            options=["Excel (.xlsx)", "CSV (.csv)"],
            horizontal=True,
            help="选择下载文件的格式"
        )
        
        excel_bytes = excel_to_bytes(result_df, "单位树结果.xlsx")
        csv_bytes = csv_to_bytes(result_df, "单位树结果.csv")
        
        download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
        
        with download_col1:
            st.markdown('<span style="font-size: 2rem;">🥔</span>', unsafe_allow_html=True)
        
        with download_col2:
            if export_format == "Excel (.xlsx)":
                st.download_button(
                    label="📥 下载结果Excel",
                    data=excel_bytes,
                    file_name=f"单位树结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            else:
                st.download_button(
                    label="📥 下载结果CSV",
                    data=csv_bytes,
                    file_name=f"单位树结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    type="primary",
                    use_container_width=True
                )
        
        with download_col3:
            st.markdown('<span style="font-size: 2rem;">🍠</span>', unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="text-align: center; color: #8B4513; margin-top: 0.5rem;">
            📊 结果：<strong>{len(result_df):,}</strong> 行 × <strong>{len(result_df.columns)}</strong> 列
        </div>
        """, unsafe_allow_html=True)
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 页面7：IP处理工具
# ============================================
def show_ip_tool():
    """显示IP处理工具"""
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🖥️ IP处理工具</h1>
        <p class="potato-subtitle">✨ IP段拆分与聚合，让IP管理更高效 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">对IP/IP段进行拆分或聚合处理，支持CIDR格式和范围格式，方便IP数据整理。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传Excel/CSV文件 → ② 字段映射 → ③ 选择处理模式 → ④ 执行处理 → ⑤ 下载结果
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 模式说明卡片
    st.markdown("""
    <div class="potato-card" style="margin-bottom: 1rem;">
        <div class="potato-card-header">📖 处理模式说明</div>
        <div style="display: flex; flex-wrap: wrap; gap: 1rem; margin-top: 0.5rem;">
            <div style="flex: 1; min-width: 280px; background: #FFF8DC; padding: 0.8rem; border-radius: 10px;">
                <div style="font-weight: 700; color: #8B4513; margin-bottom: 0.5rem;">🔀 IP段拆分</div>
                <table style="width: 100%; font-size: 0.85rem; color: #8B4513;">
                    <tr><td><b>范围格式：</b></td><td><code style="background: #FFE4C4; padding: 0.1rem 0.3rem;">192.168.1.1-192.168.1.3</code></td></tr>
                    <tr><td><b>→</b></td><td>3行：192.168.1.1, 192.168.1.2, 192.168.1.3</td></tr>
                    <tr><td><b>CIDR格式：</b></td><td><code style="background: #FFE4C4; padding: 0.1rem 0.3rem;">192.168.1.0/30</code></td></tr>
                    <tr><td><b>→</b></td><td>2行：192.168.1.1, 192.168.1.2</td></tr>
                </table>
            </div>
            <div style="flex: 1; min-width: 280px; background: #E8F5E9; padding: 0.8rem; border-radius: 10px;">
                <div style="font-weight: 700; color: #8B4513; margin-bottom: 0.5rem;">🔗 IP聚合（连续）</div>
                <table style="width: 100%; font-size: 0.85rem; color: #8B4513;">
                    <tr><td><b>输入：</b></td><td>192.168.1.1, 192.168.1.2, 192.168.1.3</td></tr>
                    <tr><td><b>→</b></td><td>全部聚合成段</td></tr>
                    <tr><td><b>输出：</b></td><td><code style="background: #C8E6C9; padding: 0.1rem 0.3rem;">192.168.1.1-192.168.1.3</code></td></tr>
                </table>
            </div>
            <div style="flex: 1; min-width: 280px; background: #FFF3E0; padding: 0.8rem; border-radius: 10px;">
                <div style="font-weight: 700; color: #8B4513; margin-bottom: 0.5rem;">🔗 IP聚合（混合）</div>
                <table style="width: 100%; font-size: 0.85rem; color: #8B4513;">
                    <tr><td><b>输入：</b></td><td>192.168.1.1, 192.168.1.2, 192.168.1.5, 192.168.1.8, 192.168.1.9</td></tr>
                    <tr><td><b>→</b></td><td>连续的成段，不连续的保留</td></tr>
                    <tr><td><b>输出：</b></td><td><code style="background: #FFE0B2; padding: 0.1rem 0.3rem;">192.168.1.1-192.168.1.2, 192.168.1.5, 192.168.1.8-192.168.1.9</code></td></tr>
                </table>
            </div>
        </div>
    </div>

    <div class="potato-card" style="margin-bottom: 1rem;">
        <div class="potato-card-header">✨ 混合格式支持（NEW）</div>
        <div style="background: #E8F4FF; padding: 1rem; border-radius: 10px; margin-top: 0.5rem;">
            <div style="color: #1565C0; font-weight: 600; margin-bottom: 0.5rem;">
                🎉 现在支持多种格式混合使用！
            </div>
            <div style="color: #0D47A1; font-size: 0.9rem; line-height: 1.8;">
                <p><b>支持格式：</b></p>
                <ul style="padding-left: 1.5rem; margin: 0.3rem 0;">
                    <li>单个IP：<code>192.168.1.1</code></li>
                    <li>IP范围：<code>192.168.1.1-192.168.1.10</code></li>
                    <li>CIDR：<code>192.168.1.0/24</code></li>
                    <li><b>混合格式：</b><code>192.168.1.1,192.168.1.5-192.168.1.10,192.168.2.0/24</code></li>
                </ul>
                <p><b>支持分隔符：</b>逗号(,)、分号(;)、中文逗号(，)、制表符、换行</p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 初始化session state
    if 'ip_df' not in st.session_state:
        st.session_state.ip_df = None
    if 'ip_result' not in st.session_state:
        st.session_state.ip_result = None
    
    # 使用说明
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0;">
            <span style="font-size: 2.5rem;">🥔</span>
            <h2 style="color: #8B4513; margin: 0.3rem 0;">使用说明</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="potato-card" style="margin-bottom: 0.8rem;">
            <div class="potato-card-header">🌱 操作步骤</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>上传 <b>Excel/CSV文件</b> 📁</li>
                <li>映射 <b>单位字段</b> 🏷️</li>
                <li>映射 <b>IP/IP段字段</b> 🖥️</li>
                <li>选择 <b>处理模式</b> ⚙️</li>
                <li>点击 <b>开始处理</b> 🚀</li>
                <li>下载 <b>结果文件</b> 📥</li>
            </ol>
        </div>
        
        <div class="potato-card">
            <div class="potato-card-header">💡 核心规则</div>
            <ul style="color: #8B4513; line-height: 1.6; font-size: 0.85rem; padding-left: 1.2rem;">
                <li>同一单位下的IP才能处理</li>
                <li>不同单位数据隔离</li>
                <li>支持异常格式跳过</li>
            </ul>
        </div>
        
        <div class="potato-card" style="margin-top: 0.8rem;">
            <div class="potato-card-header">💡 支持格式</div>
            <ul style="color: #8B4513; line-height: 1.6; font-size: 0.85rem; padding-left: 1.2rem;">
                <li>单个IP：192.168.1.1</li>
                <li>范围：192.168.1.1-192.168.1.10</li>
                <li>CIDR：192.168.1.0/24</li>
                <li><b>混合：</b>192.168.1.1,192.168.1.5-192.168.1.10,192.168.2.0/24</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 2rem;">🥔 🌿</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("🥔 IP处理工具")
    
    # 文件上传区域
    st.markdown('<div class="potato-card"><div class="potato-card-header">📁 上传数据文件</div></div>', unsafe_allow_html=True)
    
    file = st.file_uploader(
        "点击上传或拖拽Excel/CSV文件到此处",
        type=['xlsx', 'xls', 'csv'],
        help="🥔 上传包含IP数据的文件",
        key="ip_file_uploader"
    )
    
    if file:
        with st.spinner("🥔 加载中..."):
            df = load_data_file(file)
            if df is not None:
                st.session_state.ip_df = df
                st.session_state.ip_result = None
                st.markdown("""
                <div class="success-cute">✅ 文件加载成功</div>
                """, unsafe_allow_html=True)
                
                # 显示文件信息
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="potato-card"><div class="potato-card-header">📊 文件信息</div></div>', unsafe_allow_html=True)
                
                info_col1, info_col2, info_col3 = st.columns(3)
                
                with info_col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📝 总行数</div>
                        <div class="metric-value">{len(df):,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📊 总列数</div>
                        <div class="metric-value">{len(df.columns)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col3:
                    file_size_mb = file.size / (1024 * 1024)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">💾 文件大小</div>
                        <div class="metric-value">{file_size_mb:.2f} MB</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 显示所有字段
                st.markdown("**📋 可用字段：**")
                fields_display = "、".join([f"`{col}`" for col in df.columns])
                st.markdown(f"<div style='color: #8B4513;'>{fields_display}</div>", unsafe_allow_html=True)
                
                # 数据预览
                with st.expander("👁️ 预览数据（前20行）"):
                    st.dataframe(df.head(20), use_container_width=True, height=300)
    
    # 字段映射配置
    if st.session_state.ip_df is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 字段映射配置</div></div>', unsafe_allow_html=True)
        
        st.markdown("""
        <div style="background: #FFF8DC; padding: 0.8rem; border-radius: 10px; margin-bottom: 1rem;">
            <div style="color: #8B4513; font-size: 0.9rem;">
                💡 <b>单位字段</b>用于分组，同一单位的IP会在一起处理。<b>IP/IP段字段</b>包含要处理的IP数据。
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 自动识别字段
        cols = list(st.session_state.ip_df.columns)
        
        # 单位字段
        default_unit = None
        for col in cols:
            if any(keyword in col for keyword in ["单位名称", "单位", "名称", "name", "公司"]):
                default_unit = col
                break
        
        # IP字段
        default_ip = None
        for col in cols:
            if any(keyword in col.lower() for keyword in ["ip", "地址", "ip地址"]):
                default_ip = col
                break
        
        config_col1, config_col2 = st.columns(2)
        
        with config_col1:
            unit_col = st.selectbox(
                "🏷️ 单位字段（必填）",
                options=["（请选择）"] + cols,
                index=(cols.index(default_unit) + 1) if default_unit and default_unit in cols else 0,
                help="选择包含单位名称的列"
            )
            if unit_col == "（请选择）":
                unit_col = None
        
        with config_col2:
            ip_col = st.selectbox(
                "🖥️ IP/IP段字段（必填）",
                options=["（请选择）"] + cols,
                index=(cols.index(default_ip) + 1) if default_ip and default_ip in cols else 0,
                help="选择包含IP或IP段的列"
            )
            if ip_col == "（请选择）":
                ip_col = None
        
        # 字段预览
        if unit_col or ip_col:
            st.markdown("<hr>", unsafe_allow_html=True)
            
            preview_col1, preview_col2 = st.columns(2)
            
            with preview_col1:
                if unit_col:
                    st.markdown(f"**🏷️ 单位字段预览**")
                    preview_df = st.session_state.ip_df[unit_col].dropna().head(5)
                    st.write(preview_df.tolist())
                    unique_units = st.session_state.ip_df[unit_col].nunique()
                    st.caption(f"📊 共 {unique_units:,} 个唯一单位")
            
            with preview_col2:
                if ip_col:
                    st.markdown(f"**🖥️ IP字段预览**")
                    preview_df = st.session_state.ip_df[ip_col].dropna().head(5)
                    st.write(preview_df.tolist())
        
        # 处理模式选择
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 处理模式选择</div></div>', unsafe_allow_html=True)
        
        mode_col1, mode_col2, mode_col3 = st.columns(3)
        
        with mode_col1:
            st.markdown("""
            <div style="background: #FFF8DC; padding: 1rem; border-radius: 12px; text-align: center; border: 2px solid #DEB887;">
                <div style="font-size: 2rem; margin-bottom: 0.5rem;">🔀</div>
                <div style="font-weight: 700; color: #8B4513;">IP段拆分</div>
                <div style="font-size: 0.85rem; color: #D2691E; margin-top: 0.3rem;">拆分成单个IP</div>
            </div>
            """, unsafe_allow_html=True)
        
        with mode_col2:
            st.markdown("""
            <div style="background: #E8F5E9; padding: 1rem; border-radius: 12px; text-align: center; border: 2px solid #A5D6A7;">
                <div style="font-size: 2rem; margin-bottom: 0.5rem;">🔗</div>
                <div style="font-weight: 700; color: #8B4513;">IP聚合（连续）</div>
                <div style="font-size: 0.85rem; color: #D2691E; margin-top: 0.3rem;">连续的聚合成段</div>
            </div>
            """, unsafe_allow_html=True)
        
        with mode_col3:
            st.markdown("""
            <div style="background: #FFF3E0; padding: 1rem; border-radius: 12px; text-align: center; border: 2px solid #FFCC80;">
                <div style="font-size: 2rem; margin-bottom: 0.5rem;">🔗</div>
                <div style="font-weight: 700; color: #8B4513;">IP聚合（混合）</div>
                <div style="font-size: 0.85rem; color: #D2691E; margin-top: 0.3rem;">连续成段+单IP保留</div>
            </div>
            """, unsafe_allow_html=True)
        
        mode_options = {
            "IP段拆分": "split",
            "IP聚合（连续）": "aggregate_continuous",
            "IP聚合（混合）": "aggregate_mixed"
        }
        
        selected_mode = st.radio(
            "🎯 选择处理模式",
            options=list(mode_options.keys()),
            horizontal=True,
            help="选择IP处理方式"
        )
        process_mode = mode_options[selected_mode]
        
        # 模式说明
        mode_descriptions = {
            "split": """
            <div style="background: #E3F2FD; padding: 0.8rem; border-radius: 10px; margin-top: 0.5rem;">
                <div style="color: #1565C0; font-weight: 600; margin-bottom: 0.3rem;">🔀 IP段拆分</div>
                <ul style="color: #1565C0; font-size: 0.85rem; margin: 0; padding-left: 1.2rem; line-height: 1.6;">
                    <li>将IP段拆分成每个IP一行</li>
                    <li>范围格式 192.168.1.1-192.168.1.3 → 3行</li>
                    <li>CIDR格式 192.168.1.0/30 → 2行（不含网络/广播地址）</li>
                </ul>
            </div>
            """,
            "aggregate_continuous": """
            <div style="background: #E8F5E9; padding: 0.8rem; border-radius: 10px; margin-top: 0.5rem;">
                <div style="color: #2E7D32; font-weight: 600; margin-bottom: 0.3rem;">🔗 IP聚合（连续）</div>
                <ul style="color: #2E7D32; font-size: 0.85rem; margin: 0; padding-left: 1.2rem; line-height: 1.6;">
                    <li>将连续的IP聚合成IP段</li>
                    <li>输入：192.168.1.1, 192.168.1.2, 192.168.1.3</li>
                    <li>输出：192.168.1.1-192.168.1.3</li>
                </ul>
            </div>
            """,
            "aggregate_mixed": """
            <div style="background: #FFF3E0; padding: 0.8rem; border-radius: 10px; margin-top: 0.5rem;">
                <div style="color: #E65100; font-weight: 600; margin-bottom: 0.3rem;">🔗 IP聚合（混合）</div>
                <ul style="color: #E65100; font-size: 0.85rem; margin: 0; padding-left: 1.2rem; line-height: 1.6;">
                    <li>连续的IP聚合成段，不连续的保留</li>
                    <li>输入：192.168.1.1, 192.168.1.2, 192.168.1.5, 192.168.1.8, 192.168.1.9</li>
                    <li>输出：192.168.1.1-192.168.1.2, 192.168.1.5, 192.168.1.8-192.168.1.9</li>
                </ul>
            </div>
            """
        }
        
        st.markdown(mode_descriptions[process_mode], unsafe_allow_html=True)
        
        # 验证必填字段
        validation_passed = True
        if unit_col is None:
            st.markdown("""
            <div class="warning-cute" style="margin-top: 1rem;">
                ⚠️ 请选择「单位字段」🥔
            </div>
            """, unsafe_allow_html=True)
            validation_passed = False
        
        if ip_col is None:
            st.markdown("""
            <div class="warning-cute" style="margin-top: 1rem;">
                ⚠️ 请选择「IP/IP段字段」🥔
            </div>
            """, unsafe_allow_html=True)
            validation_passed = False
        
        # 执行处理按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        
        with col_btn2:
            if st.button("🚀 开始处理", type="primary", use_container_width=True):
                if unit_col is None:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择「单位字段」 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                if ip_col is None:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择「IP/IP段字段」 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                with st.spinner("🍠 正在处理IP数据..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    try:
                        start_time = time.time()
                        
                        status_text.text("🥔 正在解析数据...")
                        progress_bar.progress(0.1)
                        
                        df = st.session_state.ip_df.copy()
                        total_rows = len(df)
                        
                        # 获取其他字段（除了单位和IP字段）
                        other_cols = [col for col in df.columns if col != unit_col and col != ip_col]
                        
                        # 处理结果存储
                        results = []
                        error_stats = {}
                        total_ips = 0
                        processed_units = 0
                        
                        # 按单位分组处理
                        grouped = df.groupby(unit_col)
                        
                        for idx, (unit_name, group) in enumerate(grouped):
                            progress = 0.1 + (idx / len(grouped)) * 0.7
                            progress_bar.progress(progress)
                            status_text.text(f"🥔 处理单位 {idx + 1}/{len(grouped)}...")
                            
                            # 获取该单位的其他字段值（取第一条）
                            other_values = {}
                            for col in other_cols:
                                other_values[col] = group[col].iloc[0] if len(group) > 0 else None
                            
                            # 收集该单位的所有IP
                            unit_ips = []
                            unit_errors = []

                            for _, row in group.iterrows():
                                ip_str = row[ip_col]
                                ip_list, stats = parse_ip_range(ip_str)

                                # 统计错误信息
                                if stats["failed_segments"] > 0:
                                    for err in stats["errors"]:
                                        if err not in error_stats:
                                            error_stats[err] = 0
                                        error_stats[err] += 1

                                unit_ips.extend(ip_list)
                            
                            # 去重并统计
                            unit_ips = list(set(unit_ips))
                            total_ips += len(unit_ips)
                            
                            if process_mode == "split":
                                # IP段拆分：每个IP一行
                                for ip in sorted(unit_ips):
                                    result_row = {unit_col: unit_name, ip_col: ip}
                                    result_row.update(other_values)
                                    results.append(result_row)
                            
                            elif process_mode == "aggregate_continuous":
                                # IP聚合（连续）：全部聚合成段
                                if unit_ips:
                                    aggregated = aggregate_ips_continuous(unit_ips)
                                    result_row = {unit_col: unit_name, ip_col: ', '.join(aggregated)}
                                    result_row.update(other_values)
                                    results.append(result_row)
                            
                            else:  # aggregate_mixed
                                # IP聚合（混合）：连续的成段，不连续的保留
                                if unit_ips:
                                    aggregated = aggregate_ips_mixed(unit_ips)
                                    result_row = {unit_col: unit_name, ip_col: aggregated}
                                    result_row.update(other_values)
                                    results.append(result_row)
                            
                            processed_units += 1
                        
                        progress_bar.progress(0.85)
                        status_text.text("🍠 正在整理结果...")
                        
                        # 构建结果DataFrame
                        result_df = pd.DataFrame(results)
                        
                        # 调整列顺序
                        final_cols = [unit_col, ip_col] + other_cols
                        final_cols = [col for col in final_cols if col in result_df.columns]
                        result_df = result_df[final_cols]
                        
                        progress_bar.progress(0.95)
                        status_text.text("🍠 正在生成统计...")
                        
                        processing_time = time.time() - start_time
                        
                        # 统计信息
                        if process_mode == "split":
                            original_rows = total_rows
                            result_rows = len(result_df)
                        else:
                            original_rows = total_rows
                            result_rows = len(result_df)
                        
                        st.session_state.ip_result = {
                            'result_df': result_df,
                            'original_rows': original_rows,
                            'result_rows': result_rows,
                            'processed_units': processed_units,
                            'total_ips': total_ips,
                            'processing_time': processing_time,
                            'process_mode': selected_mode,
                            'error_stats': error_stats,
                            'unit_col': unit_col,
                            'ip_col': ip_col
                        }
                        
                        progress_bar.progress(1.0)
                        status_text.empty()
                        progress_bar.empty()
                        
                        # 显示成功消息
                        st.markdown("""
                        <div class="success-cute" style="margin-top: 1rem;">
                            🎉 IP处理完成！可以下载结果文件了 🥔🎉
                        </div>
                        """, unsafe_allow_html=True)
                        
                    except Exception as e:
                        progress_bar.empty()
                        status_text.empty()
                        st.markdown(f"""
                        <div class="error-cute">
                            ❌ 处理失败：{str(e)} 🥔
                        </div>
                        """, unsafe_allow_html=True)
    
    # 显示处理结果
    if st.session_state.ip_result:
        result = st.session_state.ip_result
        result_df = result['result_df']
        
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📊 处理结果统计</div></div>', unsafe_allow_html=True)
        
        # 统计卡片
        stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
        
        with stat_col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 原始行数</div>
                <div class="metric-value">{result['original_rows']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 处理后行数</div>
                <div class="metric-value">{result['result_rows']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">🏷️ 涉及单位</div>
                <div class="metric-value">{result['processed_units']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col4:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">🖥️ IP总数</div>
                <div class="metric-value">{result['total_ips']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        # 显示模式说明
        mode_labels = {
            "IP段拆分": "每行一个IP",
            "IP聚合（连续）": "全部聚合成段",
            "IP聚合（混合）": "连续成段+单IP"
        }
        st.markdown(f"""
        <div style="background: #FFF8DC; padding: 0.8rem; border-radius: 10px; margin-top: 0.5rem;">
            <div style="color: #8B4513; font-size: 0.9rem;">
                <b>⚙️ 处理模式：</b>{result['process_mode']} | 
                <b>效果：</b>{mode_labels.get(result['process_mode'], '')} | 
                <b>耗时：</b>{result['processing_time']:.2f}秒
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 错误统计
        error_stats = result.get('error_stats', {})
        if error_stats:
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown('<div class="potato-card"><div class="potato-card-header">⚠️ 异常IP格式（已跳过）</div></div>', unsafe_allow_html=True)
            
            error_col1, error_col2 = st.columns([2, 1])
            
            with error_col1:
                error_data = []
                for error_type, count in sorted(error_stats.items(), key=lambda x: -x[1]):
                    error_data.append({"跳过类型": error_type, "数量": count})
                
                if error_data:
                    error_df = pd.DataFrame(error_data)
                    st.dataframe(error_df, use_container_width=True, hide_index=True)
            
            with error_col2:
                st.markdown("""
                <div style="background: #FFF8DC; padding: 1rem; border-radius: 12px;">
                    <div style="color: #8B4513; font-size: 0.85rem; font-weight: 600; margin-bottom: 0.5rem;">💡 说明</div>
                    <ul style="color: #8B4513; font-size: 0.8rem; line-height: 1.6; padding-left: 1.2rem; margin: 0;">
                        <li>空值行已自动跳过</li>
                        <li>异常格式已跳过</li>
                        <li>不影响正常数据</li>
                    </ul>
                </div>
                """, unsafe_allow_html=True)
        
        # 结果预览
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">👁️ 结果预览</div></div>', unsafe_allow_html=True)
        
        # 筛选功能
        filter_col1, filter_col2 = st.columns([3, 1])
        
        with filter_col1:
            # 获取唯一单位列表
            unique_units = result_df[result['unit_col']].unique().tolist()
            filter_unit = st.multiselect(
                "🔍 按单位筛选",
                options=unique_units,
                default=[],
                help="筛选特定单位的数据"
            )
        
        with filter_col2:
            preview_limit = st.selectbox(
                "📊 预览行数",
                options=[20, 50, 100],
                index=0,
                help="选择预览行数"
            )
        
        # 应用筛选
        display_df = result_df.copy()
        if filter_unit:
            display_df = display_df[display_df[result['unit_col']].isin(filter_unit)]
        
        st.markdown(f"📊 共 **{len(display_df):,}** 条记录（原始 **{len(result_df):,}** 条）")
        
        preview_rows = min(preview_limit, len(display_df))
        st.dataframe(display_df.head(preview_rows), use_container_width=True, height=350)
        
        st.caption(f"显示前 {preview_rows} 行")
        
        # 下载按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📥 下载结果</div></div>', unsafe_allow_html=True)
        
        # 导出格式选择
        export_format = st.radio(
            "📥 选择导出格式",
            options=["Excel (.xlsx)", "CSV (.csv)"],
            horizontal=True,
            help="选择下载文件的格式"
        )
        
        excel_bytes = excel_to_bytes(result_df, "IP处理结果.xlsx")
        csv_bytes = csv_to_bytes(result_df, "IP处理结果.csv")
        
        download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
        
        with download_col1:
            st.markdown('<span style="font-size: 2rem;">🥔</span>', unsafe_allow_html=True)
        
        with download_col2:
            if export_format == "Excel (.xlsx)":
                st.download_button(
                    label="📥 下载结果Excel",
                    data=excel_bytes,
                    file_name=f"IP处理结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            else:
                st.download_button(
                    label="📥 下载结果CSV",
                    data=csv_bytes,
                    file_name=f"IP处理结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    type="primary",
                    use_container_width=True
                )
        
        with download_col3:
            st.markdown('<span style="font-size: 2rem;">🍠</span>', unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="text-align: center; color: #8B4513; margin-top: 0.5rem;">
            📊 结果：<strong>{len(result_df):,}</strong> 行 × <strong>{len(result_df.columns)}</strong> 列
        </div>
        """, unsafe_allow_html=True)
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 页面8：域名解析工具
# ============================================
def show_dns_tool():
    """显示域名解析工具"""
    # 使用内置方法，无需安装额外库
    
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🌐 域名解析工具</h1>
        <p class="potato-subtitle">✨ 查询域名的DNS解析记录 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">查询域名的DNS解析记录（CNAME/A记录），自动识别解析类型和解析值。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传Excel/CSV文件 → ② 选择域名列 → ③ 开始解析 → ④ 下载结果
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # DNS查询函数（使用内置库，无需安装额外依赖）
    def query_dns(domain):
        """查询域名DNS解析（使用系统命令，无需安装额外库）"""
        import socket
        import subprocess
        import platform
        
        try:
            domain = domain.strip()
            
            # 跳过空域名
            if not domain:
                return ('NULL', '')
            
            # 方法1：使用socket查询A记录
            try:
                ip = socket.gethostbyname(domain)
                # 检查是否有CNAME
                try:
                    # 使用nslookup查询CNAME
                    if platform.system() == 'Windows':
                        result = subprocess.run(['nslookup', '-type=CNAME', domain], 
                                              capture_output=True, text=True, timeout=10)
                        output = result.stdout
                    else:
                        result = subprocess.run(['dig', '+short', 'CNAME', domain], 
                                              capture_output=True, text=True, timeout=10)
                        output = result.stdout
                    
                    # 解析CNAME
                    if 'canonical name' in output.lower() or (output.strip() and 'dig' in str(result.args)):
                        if platform.system() == 'Windows':
                            # Windows nslookup格式
                            for line in output.split('\n'):
                                if 'canonical name' in line.lower():
                                    cname = line.split('=')[-1].strip()
                                    return ('CNAME', cname)
                        else:
                            # Linux dig格式
                            cname = output.strip().split('\n')[0]
                            if cname:
                                return ('CNAME', cname)
                except:
                    pass
                
                return ('A', ip)
            except socket.gaierror:
                return ('NULL', '域名不存在')
            except socket.timeout:
                return ('NULL', '查询超时')
            except Exception as e:
                return ('NULL', str(e)[:20])
                
        except Exception as e:
            return ('NULL', str(e)[:20])
    
    # 初始化session state
    if 'dns_df' not in st.session_state:
        st.session_state.dns_df = None
    if 'dns_result' not in st.session_state:
        st.session_state.dns_result = None
    
    # 配置区域
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div class="potato-card-header">📁 配置区域</div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns([2, 1], gap="large")
    
    with col1:
        # 上传文件
        uploaded_file = st.file_uploader(
            "📂 上传Excel或CSV文件",
            type=['xlsx', 'xls', 'csv'],
            help="支持 .xlsx, .xls, .csv 格式"
        )
    
    with col2:
        st.markdown("<div style='height: 25px;'></div>", unsafe_allow_html=True)  # 占位
        if st.button("🔄 清空数据", use_container_width=True):
            if 'dns_df' in st.session_state:
                del st.session_state.dns_df
            if 'dns_result' in st.session_state:
                del st.session_state.dns_result
            st.rerun()
    
    # 选择域名列
    if uploaded_file is not None:
        try:
            # 读取文件
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            st.session_state.dns_df = df
            
            # 显示文件信息
            st.markdown(f"""
            <div style="background: #FFF8DC; padding: 0.8rem; border-radius: 10px; margin: 0.8rem 0;">
                <span style="color: #8B4513; font-weight: 600;">📊 文件信息：</span>
                <span style="color: #D2691E;">{uploaded_file.name}</span>
                <span style="color: #8B4513;"> | 共 </span>
                <span style="color: #FF8C00; font-weight: 700;">{len(df):,}</span>
                <span style="color: #8B4513;"> 行 × </span>
                <span style="color: #FF8C00; font-weight: 700;">{len(df.columns)}</span>
                <span style="color: #8B4513;"> 列</span>
            </div>
            """, unsafe_allow_html=True)
            
            # 选择域名列
            columns = ['-- 请选择列 --'] + list(df.columns)
            selected_column = st.selectbox(
                "🔍 选择域名列",
                options=range(len(columns)),
                format_func=lambda x: columns[x],
                help="选择包含域名的列"
            )
            
            if selected_column > 0:
                domain_column = columns[selected_column]
                st.session_state.dns_domain_column = domain_column
                
                # 域名预览
                non_empty_domains = df[domain_column].dropna().astype(str).str.strip()
                non_empty_domains = non_empty_domains[non_empty_domains != '']
                
                st.markdown(f"""
                <div style="background: #FFFAF0; padding: 0.8rem; border-radius: 10px; margin: 0.8rem 0;">
                    <span style="color: #8B4513; font-weight: 600;">📋 域名预览：</span>
                    <span style="color: #8B4513;">共检测到 </span>
                    <span style="color: #32CD32; font-weight: 700;">{len(non_empty_domains):,}</span>
                    <span style="color: #8B4513;"> 个非空域名</span>
                </div>
                """, unsafe_allow_html=True)
                
                # 显示前5个域名示例
                st.markdown("<details><summary style='color:#8B4513;cursor:pointer;font-weight:600;'>👁️ 查看前5个域名示例</summary>", unsafe_allow_html=True)
                sample_domains = non_empty_domains.head(5).tolist()
                for i, d in enumerate(sample_domains, 1):
                    st.markdown(f"<div style='color:#D2691E;padding:0.2rem 0;'>• {d}</div>", unsafe_allow_html=True)
                st.markdown("</details>", unsafe_allow_html=True)
                
                # 开始解析按钮
                if st.button("🚀 开始DNS解析", type="primary", use_container_width=True):
                    with st.spinner("正在进行DNS解析，请稍候..."):
                        # 初始化进度
                        progress_bar = st.progress(0)
                        progress_text = st.empty()
                        
                        # 存储结果
                        results = []
                        total = len(df)
                        success_count = 0
                        fail_count = 0
                        
                        # 遍历每一行
                        for idx, row in df.iterrows():
                            domain = row[domain_column]
                            
                            # 更新进度
                            progress = (idx + 1) / total
                            progress_bar.progress(progress)
                            progress_text.markdown(
                                f"<span style='color:#8B4513;'>正在解析：</span>"
                                f"<span style='color:#FF8C00;font-weight:600;'>{str(domain)[:50]}</span>"
                                f"<span style='color:#8B4513;'> ({idx + 1}/{total})</span>",
                                unsafe_allow_html=True
                            )
                            
                            # 查询DNS
                            if pd.isna(domain) or str(domain).strip() == '':
                                results.append(('NULL', ''))
                            else:
                                record_type, record_value = query_dns(str(domain))
                                results.append((record_type, record_value))
                                if record_type != 'NULL':
                                    success_count += 1
                                else:
                                    fail_count += 1
                        
                        # 添加结果列
                        result_df = df.copy()
                        result_df['解析类型'] = [r[0] for r in results]
                        result_df['解析值'] = [r[1] for r in results]
                        
                        st.session_state.dns_result = result_df
                        st.session_state.dns_stats = {
                            'total': total,
                            'success': success_count,
                            'fail': fail_count
                        }
                        
                        st.rerun()
        
        except Exception as e:
            st.error(f"❌ 读取文件失败：{str(e)}")
    
    st.markdown("</div>", unsafe_allow_html=True)  # 关闭配置区域卡片
    
    # 显示结果
    if st.session_state.dns_result is not None:
        result_df = st.session_state.dns_result
        stats = st.session_state.dns_stats
        
        st.markdown("---")
        
        # 统计信息
        st.markdown("""
        <div class="potato-card" style="margin: 1rem 0;">
            <div class="potato-card-header">📊 解析结果统计</div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4, gap="medium")
        
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">总域名数</div>
                <div class="metric-value">{stats['total']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">解析成功</div>
                <div class="metric-value" style="color: #32CD32;">{stats['success']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">解析失败</div>
                <div class="metric-value" style="color: #FF6B6B;">{stats['fail']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            success_rate = (stats['success'] / stats['total'] * 100) if stats['total'] > 0 else 0
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">成功率</div>
                <div class="metric-value" style="color: #FFA500;">{success_rate:.1f}%</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("</div>", unsafe_allow_html=True)  # 关闭统计卡片
        
        # 结果预览
        st.markdown("""
        <div class="potato-card" style="margin: 1rem 0;">
            <div class="potato-card-header">👁️ 结果预览（前50条）</div>
        """, unsafe_allow_html=True)
        
        # 预览数据
        preview_df = result_df.head(50)
        
        # 根据结果添加颜色
        def highlight_result(row):
            if row['解析类型'] == 'CNAME':
                return ['background-color: #E6F3FF'] * len(row)
            elif row['解析类型'] == 'A':
                return ['background-color: #E6FFE6'] * len(row)
            else:
                return ['background-color: #FFF0F0'] * len(row)
        
        st.dataframe(
            preview_df.style.apply(highlight_result, axis=1),
            use_container_width=True,
            height=500
        )
        
        # 图例说明
        st.markdown("""
        <div style="display: flex; gap: 1rem; margin-top: 0.8rem; flex-wrap: wrap;">
            <div style="display: flex; align-items: center; gap: 0.3rem;">
                <div style="width: 20px; height: 20px; background: #E6F3FF; border-radius: 4px;"></div>
                <span style="color: #8B4513; font-size: 0.85rem;">CNAME记录</span>
            </div>
            <div style="display: flex; align-items: center; gap: 0.3rem;">
                <div style="width: 20px; height: 20px; background: #E6FFE6; border-radius: 4px;"></div>
                <span style="color: #8B4513; font-size: 0.85rem;">A记录</span>
            </div>
            <div style="display: flex; align-items: center; gap: 0.3rem;">
                <div style="width: 20px; height: 20px; background: #FFF0F0; border-radius: 4px;"></div>
                <span style="color: #8B4513; font-size: 0.85rem;">解析失败</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("</div>", unsafe_allow_html=True)  # 关闭预览卡片
        
        # 下载按钮
        st.markdown("""
        <div class="potato-card" style="margin: 1rem 0;">
            <div class="potato-card-header">💾 导出结果</div>
        """, unsafe_allow_html=True)
        
        # 生成Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='DNS解析结果', index=False)
        
        output.seek(0)
        
        col1, col2 = st.columns(2, gap="large")
        
        with col1:
            st.download_button(
                label="📥 下载Excel文件",
                data=output,
                file_name=f"DNS解析结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            # CSV下载
            csv_data = result_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 下载CSV文件",
                data=csv_data,
                file_name=f"DNS解析结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        
        st.markdown("</div>", unsafe_allow_html=True)  # 关闭下载卡片
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 页面9：数据差异行工具
# ============================================
def show_diff_tool():
    """显示数据差异行工具"""
    st.markdown("""
    <div class="potato-header">
        <h1 class="potato-title">🔍 数据差异行</h1>
        <p class="potato-subtitle">✨ 逐行比对两列数据，快速找出差异 ✨</p>
    </div>
    
    <div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>
    """, unsafe_allow_html=True)
    
    # 使用说明卡片
    st.markdown("""
    <div class="potato-card" style="margin: 1rem 0;">
        <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
            <div style="flex: 1; min-width: 250px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📖 工具用途</div>
                <div style="color: #D2691E; font-size: 0.9rem;">逐行比对同一文件中两列数据的值，快速标记并导出差异行。</div>
            </div>
            <div style="flex: 2; min-width: 300px;">
                <div style="color: #8B4513; font-weight: 600; margin-bottom: 0.5rem;">📋 使用步骤</div>
                <div style="color: #8B4513; font-size: 0.9rem;">
                    ① 上传Excel/CSV文件 → ② 选择比对列1 → ③ 选择比对列2 → ④ 执行比对 → ⑤ 查看结果与下载
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 比对规则说明
    st.markdown("""
    <div class="potato-card" style="margin-bottom: 1rem;">
        <div class="potato-card-header">📖 比对规则说明</div>
        <div style="display: flex; flex-wrap: wrap; gap: 1rem; margin-top: 0.5rem;">
            <div style="flex: 1; min-width: 280px; background: #FFF8DC; padding: 0.8rem; border-radius: 10px;">
                <div style="font-weight: 700; color: #8B4513; margin-bottom: 0.5rem;">🔍 比对逻辑</div>
                <ul style="color: #8B4513; font-size: 0.85rem; line-height: 1.8; padding-left: 1.2rem; margin: 0;">
                    <li>逐行比对两列数据的值</li>
                    <li>值不同 → 标记为差异行</li>
                    <li>自动去除首尾空格</li>
                    <li>空值视为空字符串进行比对</li>
                </ul>
            </div>
            <div style="flex: 1; min-width: 280px; background: #FFEBEE; padding: 0.8rem; border-radius: 10px;">
                <div style="font-weight: 700; color: #8B4513; margin-bottom: 0.5rem;">🎨 高亮规则</div>
                <ul style="color: #8B4513; font-size: 0.85rem; line-height: 1.8; padding-left: 1.2rem; margin: 0;">
                    <li>差异行用 <span style="background: #FFCCCC; padding: 0.1rem 0.4rem; border-radius: 4px;">浅红色背景</span> 高亮</li>
                    <li>页面展示和导出Excel均可高亮</li>
                    <li>方便快速定位差异位置</li>
                </ul>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 初始化session state
    if 'diff_df' not in st.session_state:
        st.session_state.diff_df = None
    if 'diff_result' not in st.session_state:
        st.session_state.diff_result = None
    
    # 使用说明
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0;">
            <span style="font-size: 2.5rem;">🥔</span>
            <h2 style="color: #8B4513; margin: 0.3rem 0;">使用说明</h2>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="potato-card" style="margin-bottom: 0.8rem;">
            <div class="potato-card-header">🌱 操作步骤</div>
            <ol style="color: #8B4513; line-height: 1.8; font-size: 0.9rem; padding-left: 1.2rem;">
                <li>上传 <b>Excel/CSV文件</b> 📁</li>
                <li>选择 <b>比对列1</b> 📊</li>
                <li>选择 <b>比对列2</b> 📊</li>
                <li>点击 <b>开始比对</b> 🔍</li>
                <li>查看 <b>差异结果</b> 👁️</li>
                <li>下载 <b>高亮Excel</b> 📥</li>
            </ol>
        </div>
        
        <div class="potato-card">
            <div class="potato-card-header">💡 温馨提示</div>
            <ul style="color: #8B4513; line-height: 1.7; font-size: 0.85rem; padding-left: 1.2rem;">
                <li>支持 .xlsx .xls .csv 格式</li>
                <li>CSV自动检测编码</li>
                <li>空值与任何值都视为不同</li>
                <li>导出Excel保留高亮样式</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 2rem;">🥔 🌿</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("🥔 数据差异行")
    
    # 文件上传区域
    st.markdown('<div class="potato-card"><div class="potato-card-header">📁 上传Excel/CSV文件</div></div>', unsafe_allow_html=True)
    
    file = st.file_uploader(
        "点击上传或拖拽Excel/CSV文件到此处",
        type=['xlsx', 'xls', 'csv'],
        help="🥔 上传包含要比对数据的文件",
        key="diff_file_uploader"
    )
    
    if file:
        with st.spinner("🥔 加载中..."):
            df = load_data_file(file)
            if df is not None:
                st.session_state.diff_df = df
                st.session_state.diff_result = None
                st.markdown("""
                <div class="success-cute">✅ 文件加载成功</div>
                """, unsafe_allow_html=True)
                
                # 显示文件信息
                st.markdown("<hr>", unsafe_allow_html=True)
                st.markdown('<div class="potato-card"><div class="potato-card-header">📊 文件信息</div></div>', unsafe_allow_html=True)
                
                info_col1, info_col2, info_col3 = st.columns(3)
                
                with info_col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📝 总行数</div>
                        <div class="metric-value">{len(df):,}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col2:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">📊 总列数</div>
                        <div class="metric-value">{len(df.columns)}</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with info_col3:
                    file_size_mb = file.size / (1024 * 1024)
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-label">💾 文件大小</div>
                        <div class="metric-value">{file_size_mb:.2f} MB</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # 显示所有字段
                st.markdown("**📋 可用字段：**")
                fields_display = "、".join([f"`{col}`" for col in df.columns])
                st.markdown(f"<div style='color: #8B4513;'>{fields_display}</div>", unsafe_allow_html=True)
                
                # 数据预览
                with st.expander("👁️ 预览数据（前20行）"):
                    st.dataframe(df.head(20), use_container_width=True, height=300)
    
    # 字段配置
    if st.session_state.diff_df is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">⚙️ 比对配置</div></div>', unsafe_allow_html=True)
        
        cols = list(st.session_state.diff_df.columns)
        
        config_col1, config_col2 = st.columns(2)
        
        with config_col1:
            col1 = st.selectbox(
                "📊 比对列1",
                options=["（请选择）"] + cols,
                index=0,
                help="选择要比对的第一列"
            )
            if col1 == "（请选择）":
                col1 = None
        
        with config_col2:
            col2 = st.selectbox(
                "📊 比对列2",
                options=["（请选择）"] + cols,
                index=0,
                help="选择要比对的第二列"
            )
            if col2 == "（请选择）":
                col2 = None
        
        # 字段预览
        if col1 or col2:
            st.markdown("<hr>", unsafe_allow_html=True)
            
            preview_col1, preview_col2 = st.columns(2)
            
            with preview_col1:
                if col1:
                    st.markdown(f"**📊 列1 `{col1}` 预览**")
                    preview_df = st.session_state.diff_df[col1].dropna().head(10)
                    st.write(preview_df.tolist())
                    null_count1 = st.session_state.diff_df[col1].isnull().sum()
                    st.caption(f"📊 共 {len(st.session_state.diff_df):,} 条，空值 {null_count1:,} 条")
            
            with preview_col2:
                if col2:
                    st.markdown(f"**📊 列2 `{col2}` 预览**")
                    preview_df = st.session_state.diff_df[col2].dropna().head(10)
                    st.write(preview_df.tolist())
                    null_count2 = st.session_state.diff_df[col2].isnull().sum()
                    st.caption(f"📊 共 {len(st.session_state.diff_df):,} 条，空值 {null_count2:,} 条")
        
        # 执行比对按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        
        with col_btn2:
            if st.button("🔍 开始比对", type="primary", use_container_width=True):
                if not col1:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择「比对列1」 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                if not col2:
                    st.markdown("""
                    <div class="error-cute">❌ 请选择「比对列2」 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                if col1 == col2:
                    st.markdown("""
                    <div class="error-cute">❌ 比对列1和比对列2不能相同 🥔</div>
                    """, unsafe_allow_html=True)
                    return
                
                with st.spinner("🍠 正在比对数据..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    try:
                        start_time = time.time()
                        
                        status_text.text("🥔 正在比对数据...")
                        progress_bar.progress(0.3)
                        
                        # 重置索引确保连续
                        df = st.session_state.diff_df.reset_index(drop=True).copy()
                        
                        # 验证列名是否存在
                        if col1 not in df.columns:
                            progress_bar.empty()
                            status_text.empty()
                            st.markdown(f"""
                            <div class="error-cute">❌ 列 '{col1}' 不存在，请重新选择 🥔</div>
                            """, unsafe_allow_html=True)
                            return
                        
                        if col2 not in df.columns:
                            progress_bar.empty()
                            status_text.empty()
                            st.markdown(f"""
                            <div class="error-cute">❌ 列 '{col2}' 不存在，请重新选择 🥔</div>
                            """, unsafe_allow_html=True)
                            return
                        
                        # 执行比对
                        diff_indices = compare_columns(df, col1, col2)
                        
                        progress_bar.progress(0.7)
                        status_text.text("🍠 正在整理结果...")
                        
                        # 统计信息
                        total_rows = len(df)
                        diff_count = len(diff_indices)
                        same_count = total_rows - diff_count
                        diff_rate = (diff_count / total_rows * 100) if total_rows > 0 else 0
                        
                        progress_bar.progress(1.0)
                        status_text.empty()
                        progress_bar.empty()
                        
                        processing_time = time.time() - start_time
                        
                        st.session_state.diff_result = {
                            'df': df,
                            'col1': col1,
                            'col2': col2,
                            'diff_indices': diff_indices,
                            'total_rows': total_rows,
                            'diff_count': diff_count,
                            'same_count': same_count,
                            'diff_rate': diff_rate,
                            'processing_time': processing_time
                        }
                        
                        # 显示成功消息
                        st.markdown("""
                        <div class="success-cute" style="margin-top: 1rem;">
                            🎉 比对完成！可以查看结果和下载了 🥔🎉
                        </div>
                        """, unsafe_allow_html=True)
                        
                    except Exception as e:
                        progress_bar.empty()
                        status_text.empty()
                        st.markdown(f"""
                        <div class="error-cute">
                            ❌ 比对失败：{str(e)} 🥔
                        </div>
                        """, unsafe_allow_html=True)
    
    # 显示比对结果
    if st.session_state.diff_result:
        result = st.session_state.diff_result
        df = result['df']
        diff_indices = result['diff_indices']
        
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📊 比对结果统计</div></div>', unsafe_allow_html=True)
        
        # 统计卡片
        stat_col1, stat_col2, stat_col3, stat_col4, stat_col5 = st.columns(5)
        
        with stat_col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📝 总行数</div>
                <div class="metric-value">{result['total_rows']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">✅ 相同行</div>
                <div class="metric-value" style="color: #228B22;">{result['same_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">❌ 差异行</div>
                <div class="metric-value" style="color: #FF6347;">{result['diff_count']:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col4:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">📊 差异比例</div>
                <div class="metric-value">{result['diff_rate']:.1f}%</div>
            </div>
            """, unsafe_allow_html=True)
        
        with stat_col5:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">⏱️ 处理时间</div>
                <div class="metric-value">{result['processing_time']:.2f}s</div>
            </div>
            """, unsafe_allow_html=True)
        
        # 显示比对配置
        st.markdown(f"""
        <div style="background: #FFF8DC; padding: 0.8rem; border-radius: 10px; margin-top: 0.5rem;">
            <div style="color: #8B4513; font-size: 0.9rem;">
                <b>⚙️ 比对配置：</b>
                比对列1：<code>{result['col1']}</code> | 
                比对列2：<code>{result['col2']}</code>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 差异效果提示
        if result['diff_rate'] == 0:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                🎉 太棒了！两列数据完全一致，没有差异 🥔🎉
            </div>
            """, unsafe_allow_html=True)
        elif result['diff_rate'] <= 10:
            st.markdown(f"""
            <div class="success-cute" style="margin-top: 1rem;">
                😊 不错的效果！差异比例仅 <strong>{result['diff_rate']:.1f}%</strong> 🍠
            </div>
            """, unsafe_allow_html=True)
        elif result['diff_rate'] <= 50:
            st.markdown(f"""
            <div class="warning-cute" style="margin-top: 1rem;">
                🤔 有 <strong>{result['diff_rate']:.1f}%</strong> 的行存在差异 🥔
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="warning-cute" style="margin-top: 1rem;">
                😅 差异比例较高 <strong>{result['diff_rate']:.1f}%</strong>，请检查两列数据的含义是否正确 🥔
            </div>
            """, unsafe_allow_html=True)
        
        # 结果展示
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">👁️ 结果预览（差异行高亮显示）</div></div>', unsafe_allow_html=True)
        
        # 显示模式切换
        view_mode = st.radio(
            "📋 显示模式",
            options=["显示全部数据", "仅显示差异行"],
            horizontal=True,
            help="选择要显示的数据范围"
        )
        
        # 筛选数据
        if view_mode == "仅显示差异行":
            display_df = df.iloc[diff_indices].copy().reset_index(drop=True)
            display_diff_indices = list(range(len(display_df)))  # 所有行都是差异行
        else:
            display_df = df.copy()
            display_diff_indices = diff_indices
        
        st.markdown(f"📊 共 **{len(display_df):,}** 条记录（原始 **{len(df):,}** 条）")
        
        # 使用pandas Styler进行高亮显示
        def highlight_diff_rows(row):
            """高亮差异行"""
            # 获取行的索引
            row_idx = row.name
            if row_idx in display_diff_indices:
                return ['background-color: #FFCCCC'] * len(row)
            else:
                return [''] * len(row)
        
        # 显示高亮表格
        preview_rows = min(50, len(display_df))
        preview_df = display_df.head(preview_rows)
        styled_df = preview_df.style.apply(highlight_diff_rows, axis=1)
        st.dataframe(styled_df, use_container_width=True, height=350)
        
        st.caption(f"💡 差异行已用浅红色背景高亮显示 | 显示前 {preview_rows} 行")
        
        # 下载按钮
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="potato-card"><div class="potato-card-header">📥 导出结果</div></div>', unsafe_allow_html=True)
        
        # 生成带高亮的Excel
        excel_bytes = export_with_highlight(df, diff_indices, "差异比对结果.xlsx")
        
        download_col1, download_col2, download_col3 = st.columns([1, 2, 1])
        
        with download_col1:
            st.markdown('<span style="font-size: 2rem;">🥔</span>', unsafe_allow_html=True)
        
        with download_col2:
            st.download_button(
                label="📥 下载带高亮的Excel",
                data=excel_bytes,
                file_name=f"差异比对结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        with download_col3:
            st.markdown('<span style="font-size: 2rem;">🍠</span>', unsafe_allow_html=True)
        
        st.markdown(f"""
        <div style="text-align: center; color: #8B4513; margin-top: 0.5rem;">
            📊 数据：<strong>{len(df):,}</strong> 行 × <strong>{len(df.columns)}</strong> 列 | 
            🎨 差异行：<strong>{len(diff_indices):,}</strong> 行已高亮
        </div>
        """, unsafe_allow_html=True)
    
    # 底部
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="potato-decoration">🥔 🍠 🥔 🍠 🥔</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="footer">
        <p>Made with 🥔 by 洋芋头</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
# 主应用入口
# ============================================
def main():
    """主应用入口"""
    # 侧边栏导航
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem 0 1rem 0;">
            <span style="font-size: 3rem;">🥔</span>
            <h1 style="color: #8B4513; margin: 0.3rem 0; font-size: 1.3rem;">土豆数据工具箱</h1>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        
        # 工具选项列表
        options = ["🏠 首页", "🔄 数据比对回填", "✂️ 数据拆分器", "🔗 数据聚合器", "🌐 域名提取器", "🔮 域名解析工具", "🌳 单位树构建器", "🖥️ IP处理工具", "🔍 数据差异行"]
        
        # 初始化或读取当前页面
        if 'page' not in st.session_state:
            st.session_state.page = options[0]
        
        # 根据 session_state 设置默认选中
        try:
            default_index = options.index(st.session_state.page)
        except ValueError:
            default_index = 0
        
        # 注意：不使用 key 参数，让 index 参数根据 session_state.page 生效
        page = st.radio(
            "🧭 选择工具",
            options=options,
            index=default_index,
            label_visibility="collapsed"
        )
        
        # 更新 session_state
        st.session_state.page = page
        
        st.divider()
        
        st.markdown("""
        <div style="text-align: center; padding: 0.5rem;">
            <span style="font-size: 1.5rem;">🥔 🍠 🥔</span>
            <p style="color: #8B4513; font-size: 0.85rem; margin: 0.3rem 0;">v2.6 工具箱版</p>
        </div>
        """, unsafe_allow_html=True)
    
    # 根据选择显示对应页面
    if page == "🏠 首页":
        show_home()
    elif page == "🔄 数据比对回填":
        show_compare_tool()
    elif page == "✂️ 数据拆分器":
        show_split_tool()
    elif page == "🔗 数据聚合器":
        show_aggregate_tool()
    elif page == "🌐 域名提取器":
        show_domain_tool()
    elif page == "🔮 域名解析工具":
        show_dns_tool()
    elif page == "🌳 单位树构建器":
        show_unit_tree_tool()
    elif page == "🖥️ IP处理工具":
        show_ip_tool()
    elif page == "🔍 数据差异行":
        show_diff_tool()


if __name__ == "__main__":
    main()
